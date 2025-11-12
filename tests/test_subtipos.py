"""
Testes para Subtipos de Lançamentos e integrações (CRUD, validação, filtros e exportação)
"""
from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook


def criar_subtipo(client, tipo_id: int, nome: str, ativo: bool = True):
    payload = {
        "tipo_lancamento_id": tipo_id,
        "nome": nome,
        "ativo": ativo,
    }
    resp = client.post(f"/api/tipos/{tipo_id}/subtipos", json=payload)
    assert resp.status_code == 201, resp.text
    return resp.json()


def test_crud_subtipo(client, tipo_despesa):
    # Criar
    sub = criar_subtipo(client, tipo_despesa.id, "Restaurante", True)
    assert sub["nome"] == "Restaurante"
    assert sub["ativo"] is True
    assert sub["tipo_lancamento_id"] == tipo_despesa.id

    # Listar (por tipo)
    resp = client.get(f"/api/tipos/{tipo_despesa.id}/subtipos")
    assert resp.status_code == 200
    lst = resp.json()
    assert any(s["id"] == sub["id"] for s in lst)

    # Atualizar (PATCH requer corpo completo conforme o schema atual)
    payload_update = {
        "tipo_lancamento_id": tipo_despesa.id,
        "nome": "Alimentação fora",
        "ativo": False,
    }
    resp = client.patch(f"/api/subtipos/{sub['id']}", json=payload_update)
    assert resp.status_code == 200, resp.text
    sub_upd = resp.json()
    assert sub_upd["nome"] == "Alimentação fora"
    assert sub_upd["ativo"] is False

    # Excluir
    resp = client.delete(f"/api/subtipos/{sub['id']}")
    assert resp.status_code == 200

    # Confirmar remoção
    resp = client.get(f"/api/tipos/{tipo_despesa.id}/subtipos")
    assert resp.status_code == 200
    assert all(s["id"] != sub["id"] for s in resp.json())


def test_criar_lancamento_com_subtipo(client, tipo_despesa):
    hoje = date.today()
    subtipo = criar_subtipo(client, tipo_despesa.id, "Supermercado - Alimentos")

    payload = {
        "data_lancamento": hoje.isoformat(),
        "tipo": "despesa",
        "tipo_lancamento_id": tipo_despesa.id,
        "subtipo_lancamento_id": subtipo["id"],
        "fornecedor": "Mercado Bom Preço",
        "valor_total": 150.0,
        "data_primeiro_vencimento": (hoje + timedelta(days=5)).isoformat(),
        "numero_parcelas": 1,
        "valor_medio_parcelas": 150.0,
        "observacao": "Compras do dia"
    }

    resp = client.post("/api/lancamentos", json=payload)
    assert resp.status_code == 200, resp.text
    lanc = resp.json()
    assert lanc["subtipo_lancamento_id"] == subtipo["id"]


def test_validacao_subtipo_nao_pertence(client, db_session, tipo_despesa, test_user):
    # Criar outro tipo de mesma natureza
    from app.main import TipoLancamento

    tipo_outro = TipoLancamento(
        usuario_id=test_user.id,
        nome="Transporte",
        natureza="despesa",
        created_at=date.today()
    )
    db_session.add(tipo_outro)
    db_session.commit()
    db_session.refresh(tipo_outro)

    # Criar subtipo no tipo_outro
    subtipo_outro = criar_subtipo(client, tipo_outro.id, "Combustível")

    # Tentar criar lançamento com tipo = tipo_despesa e subtipo do tipo_outro
    hoje = date.today()
    payload = {
        "data_lancamento": hoje.isoformat(),
        "tipo": "despesa",
        "tipo_lancamento_id": tipo_despesa.id,
        "subtipo_lancamento_id": subtipo_outro["id"],
        "fornecedor": "Posto XPTO",
        "valor_total": 200.0,
        "data_primeiro_vencimento": (hoje + timedelta(days=3)).isoformat(),
        "numero_parcelas": 1,
        "valor_medio_parcelas": 200.0
    }

    resp = client.post("/api/lancamentos", json=payload)
    assert resp.status_code == 400
    assert "não pertence ao tipo" in resp.json().get("detail", "")


def test_filtros_por_tipo_e_subtipo(client, db_session, test_user):
    from app.main import TipoLancamento
    hoje = date.today()

    # Criar tipos
    tipo_a = TipoLancamento(usuario_id=test_user.id, nome="Casa", natureza="despesa", created_at=hoje)
    tipo_b = TipoLancamento(usuario_id=test_user.id, nome="Lazer", natureza="despesa", created_at=hoje)
    db_session.add_all([tipo_a, tipo_b])
    db_session.commit()
    db_session.refresh(tipo_a)
    db_session.refresh(tipo_b)

    # Criar subtipos
    sub_a1 = criar_subtipo(client, tipo_a.id, "Aluguel")
    sub_b1 = criar_subtipo(client, tipo_b.id, "Cinema")

    # Lançamento 1 (tipo A, sub A1)
    resp = client.post("/api/lancamentos", json={
        "data_lancamento": hoje.isoformat(),
        "tipo": "despesa",
        "tipo_lancamento_id": tipo_a.id,
        "subtipo_lancamento_id": sub_a1["id"],
        "fornecedor": "Imobiliária XYZ",
        "valor_total": 1000.0,
        "data_primeiro_vencimento": (hoje + timedelta(days=2)).isoformat(),
        "numero_parcelas": 1,
        "valor_medio_parcelas": 1000.0
    })
    assert resp.status_code == 200

    # Lançamento 2 (tipo B, sub B1)
    resp = client.post("/api/lancamentos", json={
        "data_lancamento": hoje.isoformat(),
        "tipo": "despesa",
        "tipo_lancamento_id": tipo_b.id,
        "subtipo_lancamento_id": sub_b1["id"],
        "fornecedor": "Cinemas ABC",
        "valor_total": 50.0,
        "data_primeiro_vencimento": (hoje + timedelta(days=3)).isoformat(),
        "numero_parcelas": 1,
        "valor_medio_parcelas": 50.0
    })
    assert resp.status_code == 200

    # Filtro por tipo (tipo_a)
    resp = client.get(f"/api/lancamentos?tipo_lancamento_id={tipo_a.id}")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) >= 1
    assert all(l["tipo_lancamento_id"] == tipo_a.id for l in data)

    # Filtro por subtipo (sub_b1)
    resp = client.get(f"/api/lancamentos?subtipo_lancamento_id={sub_b1['id']}")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) >= 1
    assert all(l["subtipo_lancamento_id"] == sub_b1["id"] for l in data)


def test_excel_contem_subcategoria(client, tipo_despesa):
    hoje = date.today()
    subtipo = criar_subtipo(client, tipo_despesa.id, "Internet")

    # Criar lançamento que será exportado
    resp = client.post("/api/lancamentos", json={
        "data_lancamento": hoje.isoformat(),
        "tipo": "despesa",
        "tipo_lancamento_id": tipo_despesa.id,
        "subtipo_lancamento_id": subtipo["id"],
        "fornecedor": "Operadora NET",
        "valor_total": 120.0,
        "data_primeiro_vencimento": (hoje + timedelta(days=1)).isoformat(),
        "numero_parcelas": 1,
        "valor_medio_parcelas": 120.0
    })
    assert resp.status_code == 200

    # Exportar Excel
    resp = client.get("/api/relatorios/lancamentos-excel")
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers["content-type"]
    assert len(resp.content) > 0

    # Validar conteúdo do Excel: cabeçalhos e valor de Subcategoria
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "Subcategoria" in headers

    # Encontrar a coluna Subcategoria
    col_idx = headers.index("Subcategoria") + 1
    # Verificar se pelo menos uma linha tem o subtipo correto
    valores_sub = [ws.cell(row=r, column=col_idx).value for r in range(2, ws.max_row + 1)]
    assert any(v == "Internet" for v in valores_sub)
