from fastapi import FastAPI, Depends, HTTPException, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse, RedirectResponse
from pydantic import BaseModel, Field, ValidationError
from typing import Optional, List, Dict, Any
from sqlalchemy import Column, Integer, String, Date, Numeric, DateTime, Boolean, create_engine, func
from sqlalchemy.orm import sessionmaker, DeclarativeBase, Session
import os
import re
import smtplib
import ssl
from email.message import EmailMessage
from collections import defaultdict
import traceback
import shutil
import json
from datetime import datetime, date, timedelta
from pathlib import Path

# Importar módulos de autenticação
from app import auth
from app.auth import (
    UserCreate, UserOut, LoginRequest, Token,
    create_user, authenticate_user, get_user_by_id, get_user_by_email,
    update_user, UserUpdate, create_access_token, list_users,
    ACCESS_TOKEN_EXPIRE_MINUTES
)
from app.middleware import (
    get_current_user, get_current_admin_user, get_current_active_user,
    get_optional_user, ensure_subscription, ensure_admin_ip_allowed
)
from app.security_notifications import (
    send_login_alert_sms,
    send_login_alert_whatsapp,
    send_login_alert_email,
)

# Configuração dos caminhos
BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"
BACKUP_DIR = BASE_DIR.parent / "backups"
DATA_DIR = BASE_DIR.parent / "data"

# Criar diretórios necessários
BACKUP_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

DB_PATH = os.getenv("DB_PATH", "lancamentos.db")
# Permitir uso de DATABASE_URL (ex.: PostgreSQL) com fallback para SQLite local
DATABASE_URL = os.getenv("DATABASE_URL") or f"sqlite:///{DB_PATH}"
BRAND_DIR = STATIC_DIR / "brand"



# Garantir que o diretório do banco de dados SQLite existe
if not DATABASE_URL.startswith("postgresql") and not DATABASE_URL.startswith("mysql"):
    db_path_obj = Path(DB_PATH)
    if db_path_obj.is_absolute():
        db_dir = db_path_obj.parent
        db_dir.mkdir(parents=True, exist_ok=True)

class Base(DeclarativeBase):
    pass

# ============================================================================
# MODELO DE USUÁRIO (Autenticação)
# ============================================================================

class User(Base):
    """Modelo de usuário para autenticação"""
    __tablename__ = "users"
    
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String(255), unique=True, nullable=False, index=True)
    nome = Column(String(100), nullable=False)
    senha_hash = Column(String(255), nullable=False)
    ativo = Column(Boolean, default=True, nullable=False)
    admin = Column(Boolean, default=False, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    ultimo_acesso = Column(DateTime, nullable=True)

# ============================================================================
# COBRANÇA / ASSINATURAS
# ============================================================================

class Assinatura(Base):
    __tablename__ = "assinaturas"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, unique=True, index=True)
    status = Column(String(20), nullable=False, default="trial")  # trial | ativa | inadimplente | cancelada
    data_inicio = Column(Date, nullable=False)
    proximo_vencimento = Column(Date, nullable=True)
    valor_mensal = Column(Numeric(10,2), nullable=True)
    trial_ate = Column(Date, nullable=True)
    cancelada_em = Column(Date, nullable=True)
    created_at = Column(Date, nullable=False)

class PagamentoAssinatura(Base):
    __tablename__ = "pagamentos_assinatura"
    id = Column(Integer, primary_key=True, index=True)
    assinatura_id = Column(Integer, nullable=False, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)
    referencia = Column(String(7), nullable=False)  # YYYY-MM
    valor = Column(Numeric(10,2), nullable=False)
    data_pagamento = Column(Date, nullable=False)
    metodo = Column(String(30), nullable=True)  # pix | boleto | cartao | manual
    status = Column(String(20), nullable=False, default="confirmado")  # confirmado | pendente | falhou
    transaction_id = Column(String(100), nullable=True)

# ============================================================================
# MODELOS DE LANÇAMENTOS FINANCEIROS
# ============================================================================

class TipoLancamento(Base):
    __tablename__ = "tipos_lancamentos"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)  # FK para User
    nome = Column(String(255), nullable=False)
    natureza = Column(String(10), nullable=False)  # "despesa" | "receita"
    created_at = Column(Date, nullable=False)

class SubtipoLancamento(Base):
    __tablename__ = "subtipos_lancamentos"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)  # FK para User
    tipo_lancamento_id = Column(Integer, nullable=False, index=True)  # FK para TipoLancamento
    nome = Column(String(255), nullable=False)
    ativo = Column(Boolean, default=True, nullable=False)
    created_at = Column(Date, nullable=False)

class Lancamento(Base):
    __tablename__ = "lancamentos"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)  # FK para User
    data_lancamento = Column(Date, nullable=False)
    tipo = Column(String(10), nullable=False)  # "despesa" | "receita"
    tipo_lancamento_id = Column(Integer, nullable=True)
    subtipo_lancamento_id = Column(Integer, nullable=True, index=True)  # FK para SubtipoLancamento
    fornecedor = Column(String(255), nullable=False)
    valor_total = Column(Numeric(14,2), nullable=False)
    data_primeiro_vencimento = Column(Date, nullable=False)
    numero_parcelas = Column(Integer, nullable=False)
    valor_medio_parcelas = Column(Numeric(14,2), nullable=False)
    observacao = Column(String(1000), nullable=True)

    @property
    def tipo_lancamento(self):
        if hasattr(self, '_tipo_lancamento'):
            return self._tipo_lancamento
        return None

    @tipo_lancamento.setter
    def tipo_lancamento(self, value):
        self._tipo_lancamento = value

class Parcela(Base):
    __tablename__ = "parcelas"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)  # FK para User
    lancamento_id = Column(Integer, nullable=False)  # FK para Lancamento
    numero_parcela = Column(Integer, nullable=False)  # 1, 2, 3...
    data_vencimento = Column(Date, nullable=False)
    valor = Column(Numeric(14,2), nullable=False)
    paga = Column(Integer, nullable=False, default=0)  # 0=não paga, 1=paga
    data_pagamento = Column(Date, nullable=True)
    valor_pago = Column(Numeric(14,2), nullable=True)  # valor efetivamente pago
    forma_pagamento_id = Column(Integer, nullable=True)  # FK para FormaPagamento
    observacao_pagamento = Column(String(500), nullable=True)  # Observação do pagamento

class LancamentoRecorrente(Base):
    __tablename__ = "lancamentos_recorrentes"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)  # FK para User
    tipo = Column(String(10), nullable=False)  # "despesa" | "receita"
    tipo_lancamento_id = Column(Integer, nullable=True)
    fornecedor = Column(String(255), nullable=False)
    valor_total = Column(Numeric(14,2), nullable=False)
    dia_vencimento = Column(Integer, nullable=False)  # Dia do mês (1-31)
    numero_parcelas = Column(Integer, nullable=False)
    frequencia = Column(String(20), nullable=False)  # "mensal", "trimestral", "anual"
    ativo = Column(Integer, nullable=False, default=1)  # 0=inativo, 1=ativo
    data_inicio = Column(Date, nullable=False)  # Quando começou
    ultima_geracao = Column(Date, nullable=True)  # Última vez que foi gerado
    observacao = Column(String(1000), nullable=True)
    created_at = Column(Date, nullable=False)

class FormaPagamento(Base):
    __tablename__ = "formas_pagamento"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)  # FK para User
    nome = Column(String(100), nullable=False)  # Ex: "Conta Corrente Itaú", "Cartão Nubank"
    tipo = Column(String(20), nullable=False)  # "conta" | "cartao_credito" | "cartao_debito" | "dinheiro" | "pix"
    banco = Column(String(100), nullable=True)  # Nome do banco/instituição
    limite_credito = Column(Numeric(14,2), nullable=True)  # Apenas para cartão de crédito
    ativo = Column(Boolean, default=True, nullable=False)
    created_at = Column(Date, nullable=False)
    observacao = Column(String(500), nullable=True)

class Meta(Base):
    __tablename__ = "metas"
    id = Column(Integer, primary_key=True, index=True)
    usuario_id = Column(Integer, nullable=False, index=True)  # FK para User
    ano = Column(Integer, nullable=False)
    mes = Column(Integer, nullable=False)  # 1-12
    tipo_lancamento_id = Column(Integer, nullable=True)  # FK para TipoLancamento, null = meta geral
    valor_planejado = Column(Numeric(14,2), nullable=False)
    descricao = Column(String(500), nullable=True)
    created_at = Column(Date, nullable=False)
    updated_at = Column(Date, nullable=True)

# ============================================================================
# SEGURANÇA: RASTREAMENTO DE TENTATIVAS DE LOGIN
# ============================================================================

class LoginAttempt(Base):
    __tablename__ = "login_attempts"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String(255), nullable=False, index=True)
    user_id = Column(Integer, nullable=True, index=True)  # NULL se usuário não existir
    ip_address = Column(String(45), nullable=False)  # Suporta IPv6
    user_agent = Column(String(500), nullable=True)
    success = Column(Boolean, nullable=False, default=False)
    is_admin_attempt = Column(Boolean, nullable=False, default=False)
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)
    blocked_until = Column(DateTime, nullable=True)  # Bloqueio temporário

engine = create_engine(DATABASE_URL, echo=False, future=True)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False, future=True)
Base.metadata.create_all(bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

app = FastAPI(title="API Lançamentos", version="0.1.0")

# ================= UTILITY FUNCTIONS FOR MULTI-TENANT ISOLATION =====================
def apply_user_filter(query, model, user_id: int):
    """
    Aplica filtro de usuario_id a uma query SQLAlchemy.
    
    Args:
        query: Query SQLAlchemy
        model: Modelo da tabela (Lancamento, Parcela, FormaPagamento, etc.)
        user_id: ID do usuário atual
    
    Returns:
        Query com filtro aplicado
    """
    return query.filter(model.usuario_id == user_id)

def get_user_record(db: Session, model, record_id: int, user_id: int):
    """
    Busca um registro verificando propriedade do usuário.
    
    Args:
        db: Sessão do banco
        model: Modelo da tabela
        record_id: ID do registro
        user_id: ID do usuário atual
    
    Returns:
        Registro ou None
    """
    return db.query(model).filter(
        model.id == record_id,
        model.usuario_id == user_id
    ).first()

def get_template_context(request: Request, **kwargs) -> dict:
    """
    Prepara contexto para templates incluindo CSP nonce.
    
    Args:
        request: Request object do FastAPI
        **kwargs: Dados adicionais para o template
    
    Returns:
        Dict com contexto completo incluindo request e csp_nonce
    """
    context = {"request": request}
    # Adicionar nonce do middleware se disponível
    if hasattr(request.state, 'csp_nonce'):
        context["csp_nonce"] = request.state.csp_nonce
    # Injetar URL do logo da marca
    try:
        context["brand_logo_url"] = get_brand_logo_url()
    except Exception:
        # fallback silencioso caso algo dê errado
        context["brand_logo_url"] = "/static/icons/icon-192x192.png"
    context.update(kwargs)
    return context

# ================= BRANDING HELPERS =====================
def get_brand_logo_url() -> str:
    # Permitir sobrescrever por variável de ambiente (URL absoluta ou caminho absoluto em /static)
    env_url = os.getenv("BRAND_LOGO_URL", "").strip()
    if env_url:
        return env_url

    # Ordem de preferência de arquivos locais
    candidates = [
        STATIC_DIR / "brand" / "logo.png",
        STATIC_DIR / "brand" / "logo.svg",
        STATIC_DIR / "icons" / "domo360-logo.png",
        STATIC_DIR / "icons" / "icon-192x192.png",
    ]
    for path in candidates:
        if path.exists():
            try:
                mtime = int(path.stat().st_mtime)
            except Exception:
                mtime = 0
            rel = "/static/" + str(path.relative_to(STATIC_DIR)).replace("\\", "/")
            return f"{rel}?v={mtime}"
    # Último recurso
    return "/static/icons/icon-192x192.png"

# ================= SECURITY / RATE LIMIT / HEADERS =====================
ENVIRONMENT = os.getenv("ENVIRONMENT", "development").lower()
ALLOWED_ORIGINS_ENV = os.getenv("ALLOWED_ORIGINS", "")
if ALLOWED_ORIGINS_ENV:
    ALLOWED_ORIGINS = [o.strip() for o in ALLOWED_ORIGINS_ENV.split(",") if o.strip()]
else:
    ALLOWED_ORIGINS = [
        "http://localhost:8000", "http://127.0.0.1:8000",
        "http://localhost:8010", "http://127.0.0.1:8010"
    ]

RATE_LIMIT_WINDOW_SECONDS = 60
RATE_LIMIT_MAX_ATTEMPTS = 10
_rate_memory = defaultdict(list)
SENSITIVE_RATE_PATHS = {"/auth/login", "/auth/change-password", "/auth/register", "/admin/alterar-senha"}

def _rate_cleanup(now_ts: float, window: int, arr):
    while arr and arr[0] < now_ts - window:
        arr.pop(0)

@app.middleware("http")
async def security_headers_and_rate_limit(request: Request, call_next):
    import time
    import secrets
    path = request.url.path
    client_ip = request.client.host if request.client else "unknown"
    now_ts = time.time()

    if path in SENSITIVE_RATE_PATHS and request.method in ("POST", "PATCH"):
        bucket = _rate_memory[client_ip]
        _rate_cleanup(now_ts, RATE_LIMIT_WINDOW_SECONDS, bucket)
        bucket.append(now_ts)
        if len(bucket) > RATE_LIMIT_MAX_ATTEMPTS:
            return JSONResponse(status_code=429, content={"detail": "Muitas tentativas. Aguarde."})

    # Gerar nonce único para CSP
    nonce = secrets.token_urlsafe(16)
    request.state.csp_nonce = nonce

    response = await call_next(request)
    
    # CSP: habilita nonce e, temporariamente, 'unsafe-inline' para compatibilidade com handlers inline
    # TODO: remover 'unsafe-inline' após migrar handlers inline para addEventListener
    # Ajuste CSP: permitir carregamento de fontes Google quando usadas no login (staging)
    # Em produção pode-se optar por auto-hospedar para remover domínios externos.
    font_sources = "'self' data: https://fonts.gstatic.com"  # data: para inline SVG / PNG base64 em <img>
    style_sources = f"'self' 'unsafe-inline' https://fonts.googleapis.com"
    script_sources = f"'self' 'nonce-{nonce}' 'unsafe-inline' https://cdn.jsdelivr.net"
    csp = (
        "default-src 'self'; "
        f"script-src {script_sources}; "
        f"style-src {style_sources}; "
        f"font-src {font_sources}; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "manifest-src 'self'; "
        "frame-ancestors 'none'; "
    )
    response.headers['Content-Security-Policy'] = csp
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    if request.url.scheme == 'https' or ENVIRONMENT == 'production':
        response.headers['Strict-Transport-Security'] = 'max-age=63072000; includeSubDomains'
    return response

# Memória simples para último erro (para diagnóstico via HTTP)
LAST_ERROR: Dict[str, Any] = {}

# Handler de exceções global
@app.exception_handler(ValidationError)
async def validation_exception_handler(request: Request, exc: ValidationError):
    print(f"Erro de validação: {exc}")
    print(f"Erros: {exc.errors()}")
    # Registrar último erro
    try:
        global LAST_ERROR
        LAST_ERROR = {
            "type": "ValidationError",
            "path": str(request.url),
            "detail": exc.errors(),
            "time": datetime.utcnow().isoformat()
        }
    except Exception:
        pass
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors()}
    )

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    # Registrar último erro também para HTTPException (ex.: 400/404/500 com detail)
    try:
        global LAST_ERROR
        LAST_ERROR = {
            "type": "HTTPException",
            "path": str(request.url),
            "status": exc.status_code,
            "detail": exc.detail,
            "time": datetime.utcnow().isoformat()
        }
    except Exception:
        pass
    # Garantir resposta JSON consistente
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    if os.getenv("ENVIRONMENT", "development").lower() != "production":
        print(f"Erro geral não tratado: {type(exc).__name__}: {str(exc)}")
        print(f"Traceback: {traceback.format_exc()}")
    # Registrar último erro
    try:
        global LAST_ERROR
        LAST_ERROR = {
            "type": type(exc).__name__,
            "path": str(request.url),
            "message": str(exc),
            "trace": traceback.format_exc(),
            "time": datetime.utcnow().isoformat()
        }
    except Exception:
        pass
    return JSONResponse(
        status_code=500,
        content={"detail": "Erro interno"}
    )

# Configuração do CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PATCH", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)

# Configuração dos templates
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# Se a pasta static existir, montar os arquivos estáticos
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# Favicon handler para evitar 404 e alinhar branding
@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    brand_32 = STATIC_DIR / "icons" / "domo360-32x32.png"
    icon_32 = STATIC_DIR / "icons" / "icon-32x32.png"
    brand_192 = STATIC_DIR / "icons" / "domo360-logo.png"
    target = brand_32 if brand_32.exists() else (icon_32 if icon_32.exists() else brand_192)
    if not target.exists():
        # fallback final para o 192 padrão
        target = STATIC_DIR / "icons" / "icon-192x192.png"
    return FileResponse(str(target), media_type="image/png")


# ========== MIDDLEWARE: BLOQUEIO POR ASSINATURA (somente usuários autenticados) ==========

@app.middleware("http")
async def billing_subscription_guard(request: Request, call_next):
    try:
        # Métodos que modificam estado
        if request.method in ("POST", "PUT", "PATCH", "DELETE"):
            path = request.url.path
            # Rotas liberadas
            allowed_prefixes = (
                "/auth", "/api/billing", "/api/health", "/health", "/api/debug",
                "/static", "/offline", "/sw.js"
            )
            if path.startswith(allowed_prefixes):
                return await call_next(request)

            # Tentar obter token do header ou cookie
            token = None
            auth_header = request.headers.get("Authorization")
            if auth_header and auth_header.lower().startswith("bearer "):
                token = auth_header.split(" ", 1)[1]
            elif "access_token" in request.cookies:
                token = request.cookies.get("access_token")

            if token:
                # Validar token e checar assinatura
                token_data = auth.decode_access_token(token)
                if token_data and token_data.user_id:
                    db = SessionLocal()
                    try:
                        user = get_user_by_id(db, token_data.user_id)
                        if user and user.ativo:
                            sub = db.query(Assinatura).filter(Assinatura.usuario_id == user.id).first()
                            hoje = date.today()
                            if not sub:
                                # cria trial automática de 14 dias
                                trial_ate = hoje + timedelta(days=14)
                                sub = Assinatura(
                                    usuario_id=user.id,
                                    status="trial",
                                    data_inicio=hoje,
                                    proximo_vencimento=trial_ate,
                                    valor_mensal=None,
                                    trial_ate=trial_ate,
                                    created_at=hoje
                                )
                                db.add(sub)
                                db.commit()
                            else:
                                # Bloquear se vencida e não cancelada
                                if sub.proximo_vencimento and sub.proximo_vencimento < hoje and sub.status != "cancelada":
                                    if sub.status != "inadimplente":
                                        sub.status = "inadimplente"
                                        db.commit()
                                    return JSONResponse(
                                        status_code=402,
                                        content={
                                            "detail": {
                                                "message": "Assinatura vencida. Regularize o pagamento para continuar.",
                                                "status": sub.status,
                                                "proximo_vencimento": sub.proximo_vencimento.isoformat() if sub.proximo_vencimento else None,
                                                "trial_ate": sub.trial_ate.isoformat() if sub.trial_ate else None
                                            }
                                        }
                                    )
                    finally:
                        db.close()
            # Se não houver token, não bloqueia (modo legado)
        return await call_next(request)
    except Exception as e:
        # Em caso de erro no guard, não derruba a app, apenas prossegue
        return await call_next(request)


# ========== MIDDLEWARE: VERIFICAÇÃO DE ORIGIN/REFERER PARA MÉTODOS DE ESCRITA ==========

@app.middleware("http")
async def csrf_origin_guard(request: Request, call_next):
    try:
        if request.method in ("POST", "PUT", "PATCH", "DELETE"):
            path = request.url.path
            # Prefixos liberados
            allowed_prefixes = ("/static", "/api/health", "/api/debug", "/sw.js", "/offline")
            if path.startswith(allowed_prefixes):
                return await call_next(request)

            # Apenas verificar quando há cookie de sessão (autenticação por cookie)
            if "access_token" in request.cookies:
                expected_origin = f"{request.url.scheme}://{request.url.netloc}"
                origin = request.headers.get("Origin")
                referer = request.headers.get("Referer")
                origin_ok = False

                if origin:
                    if origin == expected_origin or origin in ALLOWED_ORIGINS:
                        origin_ok = True
                else:
                    # Sem Origin: usar Referer como fallback
                    if referer and referer.startswith(expected_origin + "/"):
                        origin_ok = True

                if not origin_ok:
                    return JSONResponse(status_code=403, content={"detail": "Origem não permitida"})

        return await call_next(request)
    except Exception:
        return await call_next(request)


# ========== SAÚDE DA APLICAÇÃO ==========
@app.get("/api/health")
async def health_check():
    """Endpoint simples para verificação de disponibilidade da API"""
    try:
        return {"status": "ok", "time": datetime.utcnow().isoformat()}
    except Exception as e:
        # Não deve acontecer, mas evita quebrar health
        return JSONResponse(status_code=500, content={"status": "error", "detail": str(e)})

@app.get("/api/debug/last-error")
async def get_last_error():
    """Retorna informações do último erro registrado pelo servidor (apenas para diagnóstico local)."""
    try:
        return LAST_ERROR or {"detail": "sem erros registrados nesta sessão"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ========== FUNÇÕES DE BACKUP E VALIDAÇÃO ==========

def criar_backup() -> Dict[str, Any]:
    """Cria backup do banco de dados SQLite com timestamp"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{timestamp}.db"
        backup_path = BACKUP_DIR / backup_filename
        
        # Copiar arquivo do banco
        db_source = Path(DB_PATH)
        if not db_source.exists():
            raise FileNotFoundError(f"Banco de dados não encontrado: {DB_PATH}")
        
        shutil.copy2(db_source, backup_path)
        
        # Obter tamanho do arquivo
        file_size = backup_path.stat().st_size
        
        return {
            "success": True,
            "filename": backup_filename,
            "path": str(backup_path),
            "size": file_size,
            "size_mb": round(file_size / (1024 * 1024), 2),
            "timestamp": timestamp,
            "created_at": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

def listar_backups() -> List[Dict[str, Any]]:
    """Lista todos os backups disponíveis"""
    backups = []
    
    for backup_file in sorted(BACKUP_DIR.glob("backup_*.db"), reverse=True):
        stat = backup_file.stat()
        
        # Extrair timestamp do nome do arquivo
        filename = backup_file.name
        timestamp_str = filename.replace("backup_", "").replace(".db", "")
        
        try:
            created_date = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
        except:
            created_date = datetime.fromtimestamp(stat.st_mtime)
        
        backups.append({
            "filename": filename,
            "path": str(backup_file),
            "size": stat.st_size,
            "size_mb": round(stat.st_size / (1024 * 1024), 2),
            "created_at": created_date.isoformat(),
            "age_days": (datetime.now() - created_date).days
        })
    
    return backups

def limpar_backups_antigos(dias: int = 30):
    """Remove backups mais antigos que N dias"""
    limite = datetime.now() - timedelta(days=dias)
    removidos = []
    
    for backup_file in BACKUP_DIR.glob("backup_*.db"):
        stat = backup_file.stat()
        created = datetime.fromtimestamp(stat.st_mtime)
        
        if created < limite:
            try:
                backup_file.unlink()
                removidos.append(backup_file.name)
            except Exception as e:
                print(f"Erro ao remover backup {backup_file.name}: {e}")
    
    return removidos

def restaurar_backup(backup_filename: str) -> Dict[str, Any]:
    """Restaura banco de dados de um backup"""
    try:
        backup_path = BACKUP_DIR / backup_filename
        
        if not backup_path.exists():
            raise FileNotFoundError(f"Backup não encontrado: {backup_filename}")
        
        # Criar backup do estado atual antes de restaurar
        backup_atual = criar_backup()
        
        # Restaurar
        db_dest = Path(DB_PATH)
        shutil.copy2(backup_path, db_dest)
        
        return {
            "success": True,
            "restored_from": backup_filename,
            "backup_created": backup_atual.get("filename"),
            "message": "Banco restaurado com sucesso. Backup do estado anterior criado."
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

def exportar_dados_json(db: Session) -> Dict[str, Any]:
    """Exporta todos os dados do banco em formato JSON"""
    try:
        # Buscar todos os tipos
        tipos = db.query(TipoLancamento).all()
        tipos_data = [{
            "id": t.id,
            "nome": t.nome,
            "natureza": t.natureza,
            "created_at": t.created_at.isoformat() if t.created_at else None
        } for t in tipos]
        
        # Buscar todos os lançamentos
        lancamentos = db.query(Lancamento).all()
        lancamentos_data = [{
            "id": l.id,
            "data_lancamento": l.data_lancamento.isoformat() if l.data_lancamento else None,
            "tipo": l.tipo,
            "tipo_lancamento_id": l.tipo_lancamento_id,
            "fornecedor": l.fornecedor,
            "valor_total": float(l.valor_total),
            "data_primeiro_vencimento": l.data_primeiro_vencimento.isoformat() if l.data_primeiro_vencimento else None,
            "numero_parcelas": l.numero_parcelas,
            "valor_medio_parcelas": float(l.valor_medio_parcelas),
            "observacao": l.observacao
        } for l in lancamentos]
        
        # Buscar todas as parcelas
        parcelas = db.query(Parcela).all()
        parcelas_data = [{
            "id": p.id,
            "lancamento_id": p.lancamento_id,
            "numero_parcela": p.numero_parcela,
            "data_vencimento": p.data_vencimento.isoformat() if p.data_vencimento else None,
            "valor": float(p.valor),
            "paga": p.paga,
            "data_pagamento": p.data_pagamento.isoformat() if p.data_pagamento else None,
            "valor_pago": float(p.valor_pago) if p.valor_pago else None
        } for p in parcelas]
        
        # Buscar lançamentos recorrentes
        recorrentes = db.query(LancamentoRecorrente).all()
        recorrentes_data = [{
            "id": r.id,
            "tipo": r.tipo,
            "tipo_lancamento_id": r.tipo_lancamento_id,
            "fornecedor": r.fornecedor,
            "valor_total": float(r.valor_total),
            "dia_vencimento": r.dia_vencimento,
            "numero_parcelas": r.numero_parcelas,
            "frequencia": r.frequencia,
            "ativo": r.ativo,
            "data_inicio": r.data_inicio.isoformat() if r.data_inicio else None,
            "ultima_geracao": r.ultima_geracao.isoformat() if r.ultima_geracao else None,
            "observacao": r.observacao,
            "created_at": r.created_at.isoformat() if r.created_at else None
        } for r in recorrentes]
        
        export_data = {
            "export_date": datetime.now().isoformat(),
            "version": "1.0",
            "tipos_lancamentos": tipos_data,
            "lancamentos": lancamentos_data,
            "parcelas": parcelas_data,
            "lancamentos_recorrentes": recorrentes_data,
            "stats": {
                "total_tipos": len(tipos_data),
                "total_lancamentos": len(lancamentos_data),
                "total_parcelas": len(parcelas_data),
                "total_recorrentes": len(recorrentes_data)
            }
        }
        
        return export_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar dados: {str(e)}")

def validar_integridade(db: Session) -> Dict[str, Any]:
    """Valida integridade dos dados no banco"""
    problemas = []
    
    try:
        # 1. Verificar parcelas órfãs (sem lançamento)
        parcelas_orfas = db.query(Parcela).filter(
            ~Parcela.lancamento_id.in_(db.query(Lancamento.id))
        ).count()
        
        if parcelas_orfas > 0:
            problemas.append({
                "tipo": "parcelas_orfas",
                "severidade": "alta",
                "quantidade": parcelas_orfas,
                "descricao": f"{parcelas_orfas} parcelas sem lançamento correspondente"
            })
        
        # 2. Verificar lançamentos com valores negativos
        lancamentos_negativos = db.query(Lancamento).filter(
            Lancamento.valor_total < 0
        ).count()
        
        if lancamentos_negativos > 0:
            problemas.append({
                "tipo": "valores_negativos",
                "severidade": "media",
                "quantidade": lancamentos_negativos,
                "descricao": f"{lancamentos_negativos} lançamentos com valor negativo"
            })
        
        # 3. Verificar inconsistência entre valor_total e soma das parcelas
        lancamentos = db.query(Lancamento).all()
        inconsistencias = 0
        
        for lanc in lancamentos:
            parcelas = db.query(Parcela).filter(Parcela.lancamento_id == lanc.id).all()
            soma_parcelas = sum(float(p.valor) for p in parcelas)
            
            # Tolerar diferença de até 0.01 por arredondamento
            if abs(float(lanc.valor_total) - soma_parcelas) > 0.01:
                inconsistencias += 1
        
        if inconsistencias > 0:
            problemas.append({
                "tipo": "inconsistencia_valores",
                "severidade": "alta",
                "quantidade": inconsistencias,
                "descricao": f"{inconsistencias} lançamentos com diferença entre valor_total e soma das parcelas"
            })
        
        # 4. Verificar parcelas pagas sem data ou valor de pagamento
        parcelas_pagas_invalidas = db.query(Parcela).filter(
            Parcela.paga == 1,
            (Parcela.data_pagamento == None) | (Parcela.valor_pago == None)
        ).count()
        
        if parcelas_pagas_invalidas > 0:
            problemas.append({
                "tipo": "parcelas_pagas_invalidas",
                "severidade": "media",
                "quantidade": parcelas_pagas_invalidas,
                "descricao": f"{parcelas_pagas_invalidas} parcelas marcadas como pagas sem data/valor de pagamento"
            })
        
        # 5. Verificar tipos de lançamento órfãos
        tipos_orfaos = db.query(TipoLancamento).filter(
            ~TipoLancamento.id.in_(db.query(Lancamento.tipo_lancamento_id).distinct())
        ).count()
        
        return {
            "data_verificacao": datetime.now().isoformat(),
            "status": "ok" if len(problemas) == 0 else "problemas_encontrados",
            "total_problemas": len(problemas),
            "problemas": problemas,
            "estatisticas": {
                "total_lancamentos": db.query(Lancamento).count(),
                "total_parcelas": db.query(Parcela).count(),
                "total_tipos": db.query(TipoLancamento).count(),
                "parcelas_pagas": db.query(Parcela).filter(Parcela.paga == 1).count(),
                "parcelas_pendentes": db.query(Parcela).filter(Parcela.paga == 0).count()
            }
        }
    except Exception as e:
        return {
            "status": "erro",
            "error": str(e)
        }


class LancamentoIn(BaseModel):
    data_lancamento: str = Field(..., description="YYYY-MM-DD")
    tipo: str = Field(..., pattern="^(despesa|receita)$")
    tipo_lancamento_id: Optional[int] = None
    subtipo_lancamento_id: Optional[int] = None
    fornecedor: str = Field(..., min_length=1, max_length=255)
    valor_total: float = Field(..., gt=0)
    data_primeiro_vencimento: str = Field(..., description="YYYY-MM-DD")
    numero_parcelas: int = Field(..., gt=0)
    valor_medio_parcelas: float = Field(..., gt=0)
    observacao: Optional[str] = None

class TipoLancamentoIn(BaseModel):
    nome: str = Field(..., min_length=1, max_length=255)
    natureza: str = Field(..., pattern="^(despesa|receita)$")

class TipoLancamentoOut(TipoLancamentoIn):
    id: int
    created_at: str

    model_config = {
        "from_attributes": True
    }

    @classmethod
    def model_validate(cls, obj, *args, **kwargs):
        data = {
            'id': obj.id,
            'nome': obj.nome,
            'natureza': obj.natureza,
            'created_at': obj.created_at.isoformat() if obj.created_at else None
        }
        return cls(**data)

class SubtipoLancamentoIn(BaseModel):
    tipo_lancamento_id: int = Field(..., gt=0)
    nome: str = Field(..., min_length=1, max_length=255)
    ativo: bool = True

class SubtipoLancamentoOut(BaseModel):
    id: int
    tipo_lancamento_id: int
    nome: str
    ativo: bool
    created_at: str

    model_config = {
        "from_attributes": True
    }

    @classmethod
    def model_validate(cls, obj, *args, **kwargs):
        data = {
            'id': obj.id,
            'tipo_lancamento_id': obj.tipo_lancamento_id,
            'nome': obj.nome,
            'ativo': obj.ativo,
            'created_at': obj.created_at.isoformat() if obj.created_at else None
        }
        return cls(**data)

class FormaPagamentoIn(BaseModel):
    nome: str = Field(..., min_length=1, max_length=100)
    tipo: str = Field(..., pattern="^(conta|cartao_credito|cartao_debito|dinheiro|pix)$")
    banco: Optional[str] = Field(None, max_length=100)
    limite_credito: Optional[float] = Field(None, gt=0)
    ativo: bool = True
    observacao: Optional[str] = Field(None, max_length=500)

class FormaPagamentoOut(BaseModel):
    id: int
    nome: str
    tipo: str
    banco: Optional[str]
    limite_credito: Optional[float]
    ativo: bool
    created_at: str
    observacao: Optional[str]
    
    model_config = {
        "from_attributes": True
    }
    
    @classmethod
    def from_orm(cls, obj):
        return cls(
            id=obj.id,
            nome=obj.nome,
            tipo=obj.tipo,
            banco=obj.banco,
            limite_credito=float(obj.limite_credito) if obj.limite_credito else None,
            ativo=obj.ativo,
            created_at=obj.created_at.isoformat() if hasattr(obj.created_at, 'isoformat') else obj.created_at,
            observacao=obj.observacao
        )

class ParcelaOut(BaseModel):
    id: int
    lancamento_id: int
    numero_parcela: int
    data_vencimento: str
    valor: float
    paga: bool
    data_pagamento: Optional[str]
    valor_pago: Optional[float]
    forma_pagamento_id: Optional[int]
    observacao_pagamento: Optional[str]
    
    model_config = {
        "from_attributes": True
    }
    
    @classmethod
    def from_orm(cls, obj):
        return cls(
            id=obj.id,
            lancamento_id=obj.lancamento_id,
            numero_parcela=obj.numero_parcela,
            data_vencimento=obj.data_vencimento.isoformat() if hasattr(obj.data_vencimento, 'isoformat') else obj.data_vencimento,
            valor=float(obj.valor),
            paga=bool(obj.paga),
            data_pagamento=obj.data_pagamento.isoformat() if obj.data_pagamento and hasattr(obj.data_pagamento, 'isoformat') else obj.data_pagamento,
            valor_pago=float(obj.valor_pago) if obj.valor_pago else None,
            forma_pagamento_id=obj.forma_pagamento_id if hasattr(obj, 'forma_pagamento_id') else None,
            observacao_pagamento=obj.observacao_pagamento if hasattr(obj, 'observacao_pagamento') else None
        )

class LancamentoOut(BaseModel):
    id: int
    data_lancamento: str
    tipo: str
    tipo_lancamento_id: Optional[int]
    subtipo_lancamento_id: Optional[int]
    fornecedor: str
    valor_total: float
    data_primeiro_vencimento: str
    numero_parcelas: int
    valor_medio_parcelas: float
    observacao: Optional[str]
    parcelas: Optional[List['ParcelaOut']] = []
    
    model_config = {
        "from_attributes": True
    }

    @classmethod
    def from_orm(cls, obj, incluir_parcelas=False):
        data = {
            'id': obj.id,
            'data_lancamento': obj.data_lancamento.isoformat() if hasattr(obj.data_lancamento, 'isoformat') else obj.data_lancamento,
            'tipo': obj.tipo,
            'tipo_lancamento_id': obj.tipo_lancamento_id,
            'subtipo_lancamento_id': obj.subtipo_lancamento_id,
            'fornecedor': obj.fornecedor,
            'valor_total': float(obj.valor_total),
            'data_primeiro_vencimento': obj.data_primeiro_vencimento.isoformat() if hasattr(obj.data_primeiro_vencimento, 'isoformat') else obj.data_primeiro_vencimento,
            'numero_parcelas': obj.numero_parcelas,
            'valor_medio_parcelas': float(obj.valor_medio_parcelas),
            'observacao': obj.observacao,
            'parcelas': []
        }
        
        if incluir_parcelas and hasattr(obj, '_parcelas') and obj._parcelas:
            data['parcelas'] = [ParcelaOut.from_orm(p) for p in obj._parcelas]
        
        return cls(**data)

# ======================
# ENDPOINTS DE AUTENTICAÇÃO
# ======================

@app.post("/auth/register", response_model=UserOut, status_code=201)
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    """
    Registra um novo usuário.
    """
    try:
        # Verifica se email já existe
        existing_user = get_user_by_email(db, user_data.email)
        if existing_user:
            raise HTTPException(status_code=400, detail="Email já cadastrado")
        
        # Cria novo usuário
        new_user = create_user(db, user_data)
        return new_user
    except ValueError as e:
        # Erros de validação de senha
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        # Log do erro real para debug
        print(f"Erro ao criar usuário: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail="Erro ao criar conta. Tente novamente.")

@app.post("/auth/login", response_model=Token)
async def login(
    response: Response,
    login_data: LoginRequest,
    db: Session = Depends(get_db)
):
    """
    Autentica usuário e retorna token JWT.
    """
    # Autentica usuário
    user = authenticate_user(db, login_data.email, login_data.password)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Email ou senha incorretos"
        )
    
    if not user.ativo:
        raise HTTPException(
            status_code=403,
            detail="Usuário inativo"
        )
    
    # Cria token de acesso com id e email
    access_token = create_access_token(data={"sub": str(user.id), "email": user.email})
    
    # Define cookie com o token (httponly para segurança)
    cookie_params = {
        "key": "access_token",
        "value": access_token,
        "httponly": True,
        "max_age": ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        "samesite": "strict" if ENVIRONMENT == 'production' else 'lax'
    }
    if ENVIRONMENT == 'production':
        cookie_params["secure"] = True
    response.set_cookie(**cookie_params)
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": UserOut.from_orm(user)
    }

@app.post("/auth/logout")
async def logout(response: Response):
    """
    Faz logout do usuário removendo o cookie.
    """
    response.delete_cookie(key="access_token")
    return {"message": "Logout realizado com sucesso"}

@app.get("/auth/me", response_model=UserOut)
async def get_me(current_user: User = Depends(get_current_active_user)):
    """
    Retorna informações do usuário autenticado.
    """
    return current_user

@app.patch("/auth/me", response_model=UserOut)
async def update_me(
    user_update: UserUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Atualiza informações do usuário autenticado.
    """
    updated_user = update_user(db, current_user.id, user_update)
    return updated_user

class ChangePasswordRequest(BaseModel):
    senha_atual: str = Field(..., min_length=1)
    senha_nova: str = Field(..., min_length=8, max_length=100)
    senha_nova_confirmacao: str = Field(..., min_length=8, max_length=100)

@app.post("/auth/change-password")
async def change_password(
    password_data: ChangePasswordRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Altera a senha do usuário autenticado.
    """
    from app.auth import verify_password, get_password_hash, validate_password_strength
    
    # Validar que as novas senhas são iguais
    if password_data.senha_nova != password_data.senha_nova_confirmacao:
        raise HTTPException(
            status_code=400,
            detail="As senhas novas não conferem"
        )
    
    # Verificar senha atual
    if not verify_password(password_data.senha_atual, current_user.senha_hash):
        raise HTTPException(
            status_code=400,
            detail="Senha atual incorreta"
        )
    
    # CORREÇÃO: Buscar o usuário diretamente da sessão do banco
    # para garantir que a modificação seja persistida
    user_in_db = db.query(User).filter(User.id == current_user.id).first()
    if not user_in_db:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    if not validate_password_strength(password_data.senha_nova):
        raise HTTPException(status_code=400, detail="Senha não atende aos requisitos mínimos")
    new_hash = get_password_hash(password_data.senha_nova)
    user_in_db.senha_hash = new_hash
    
    try:
        db.commit()
        db.refresh(user_in_db)  # Garantir que pegamos o valor atualizado
    except Exception as e:
        db.rollback()
        print(f"[ERRO change_password] Falha ao commitar nova senha do usuario {current_user.id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao salvar nova senha")

    # Debug: validar imediatamente se o hash confere
    try:
        ok_verify = verify_password(password_data.senha_nova, user_in_db.senha_hash)
    except Exception as e:
        ok_verify = False
        print(f"[ERRO change_password] Falha ao verificar hash recém criado: {e}")

    if not ok_verify:
        raise HTTPException(status_code=500, detail="Falha ao validar nova senha")

    return {"message": "Senha alterada com sucesso"}

@app.get("/auth/users", response_model=List[UserOut])
async def list_all_users(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """
    Lista todos os usuários (somente admin).
    """
    users = list_users(db, skip=skip, limit=limit)
    return users

# ======================
# BILLING / ASSINATURAS
# ======================

class AssinaturaOut(BaseModel):
    status: str
    data_inicio: Optional[str] = None
    proximo_vencimento: Optional[str] = None
    valor_mensal: Optional[float] = None
    trial_ate: Optional[str] = None

    @classmethod
    def from_orm(cls, obj):
        return cls(
            status=obj.status,
            data_inicio=obj.data_inicio.isoformat() if obj.data_inicio else None,
            proximo_vencimento=obj.proximo_vencimento.isoformat() if obj.proximo_vencimento else None,
            valor_mensal=float(obj.valor_mensal) if obj.valor_mensal is not None else None,
            trial_ate=obj.trial_ate.isoformat() if obj.trial_ate else None
        )

class AssinaturaStartIn(BaseModel):
    valor_mensal: Optional[float] = Field(None, gt=0)
    trial_dias: Optional[int] = Field(None, ge=0, le=60)

class PagamentoIn(BaseModel):
    valor: float = Field(..., gt=0)
    referencia: Optional[str] = Field(None, pattern=r"^\d{4}-\d{2}$")  # YYYY-MM
    metodo: Optional[str] = Field("manual", pattern=r"^(manual|pix|boleto|cartao)$")
    transaction_id: Optional[str] = None

class PagamentoOut(BaseModel):
    id: int
    referencia: str
    valor: float
    data_pagamento: str
    metodo: Optional[str]
    status: str
    transaction_id: Optional[str]

    @classmethod
    def from_orm(cls, obj):
        return cls(
            id=obj.id,
            referencia=obj.referencia,
            valor=float(obj.valor),
            data_pagamento=obj.data_pagamento.isoformat(),
            metodo=obj.metodo,
            status=obj.status,
            transaction_id=obj.transaction_id
        )

@app.get("/api/billing/assinatura", response_model=AssinaturaOut)
async def obter_minha_assinatura(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    sub = db.query(Assinatura).filter(Assinatura.usuario_id == current_user.id).first()
    if not sub:
        # Aproveita a lógica de ensure_subscription para criar trial
        await ensure_subscription(current_user, db)  # type: ignore
        sub = db.query(Assinatura).filter(Assinatura.usuario_id == current_user.id).first()
    return AssinaturaOut.from_orm(sub)

@app.post("/api/billing/assinatura/start", response_model=AssinaturaOut)
async def iniciar_assinatura(
    dados: AssinaturaStartIn,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    from dateutil.relativedelta import relativedelta
    hoje = date.today()
    sub = db.query(Assinatura).filter(Assinatura.usuario_id == current_user.id).first()
    if not sub:
        sub = Assinatura(
            usuario_id=current_user.id,
            status="ativa",
            data_inicio=hoje,
            proximo_vencimento=hoje + relativedelta(months=1),
            valor_mensal=("{:.2f}".format(dados.valor_mensal) if dados.valor_mensal else None),
            trial_ate=(hoje + timedelta(days=dados.trial_dias)) if dados.trial_dias else None,
            created_at=hoje
        )
        db.add(sub)
    else:
        sub.status = "ativa"
        sub.valor_mensal = ("{:.2f}".format(dados.valor_mensal) if dados.valor_mensal else sub.valor_mensal)
        if not sub.proximo_vencimento or sub.proximo_vencimento < hoje:
            sub.proximo_vencimento = hoje + relativedelta(months=1)
    db.commit()
    db.refresh(sub)
    return AssinaturaOut.from_orm(sub)

@app.post("/api/billing/pagamentos", response_model=PagamentoOut)
async def registrar_pagamento(
    dados: PagamentoIn,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    from dateutil.relativedelta import relativedelta
    # Garantir assinatura existente
    sub = db.query(Assinatura).filter(Assinatura.usuario_id == current_user.id).first()
    if not sub:
        await ensure_subscription(current_user, db)  # type: ignore
        sub = db.query(Assinatura).filter(Assinatura.usuario_id == current_user.id).first()

    hoje = date.today()
    # Calcular referência padrão YYYY-MM
    if not dados.referencia:
        dados.referencia = hoje.strftime("%Y-%m")

    pagamento = PagamentoAssinatura(
        assinatura_id=sub.id,
        usuario_id=current_user.id,
        referencia=dados.referencia,
        valor="{:.2f}".format(dados.valor),
        data_pagamento=hoje,
        metodo=dados.metodo,
        status="confirmado",
        transaction_id=dados.transaction_id
    )
    db.add(pagamento)

    # Atualizar assinatura
    sub.status = "ativa"
    if sub.proximo_vencimento and sub.proximo_vencimento >= hoje:
        sub.proximo_vencimento = sub.proximo_vencimento + relativedelta(months=1)
    else:
        sub.proximo_vencimento = hoje + relativedelta(months=1)

    db.commit()
    db.refresh(pagamento)
    return PagamentoOut.from_orm(pagamento)

@app.get("/api/billing/pagamentos", response_model=List[PagamentoOut])
async def listar_pagamentos(
    limit: int = 12,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    pagamentos = db.query(PagamentoAssinatura).filter(
        PagamentoAssinatura.usuario_id == current_user.id
    ).order_by(PagamentoAssinatura.data_pagamento.desc()).limit(limit).all()
    return [PagamentoOut.from_orm(p) for p in pagamentos]

@app.get("/api/admin/billing/stats")
async def admin_billing_stats(
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    hoje = date.today()
    total_usuarios = db.query(User).count()

    # Sub em dia: proximo_vencimento >= hoje e status != cancelada
    pagos = db.query(Assinatura).filter(
        Assinatura.proximo_vencimento != None,
        Assinatura.proximo_vencimento >= hoje,
        Assinatura.status != "cancelada"
    ).count()

    # Inadimplentes: proximo_vencimento < hoje e não cancelada
    vencidos = db.query(Assinatura).filter(
        Assinatura.proximo_vencimento != None,
        Assinatura.proximo_vencimento < hoje,
        Assinatura.status != "cancelada"
    ).count()

    # Utilizando: último acesso em 7 dias
    from datetime import timedelta as dt_timedelta
    ativos = db.query(User).filter(
        User.ultimo_acesso != None,
        User.ultimo_acesso >= datetime.utcnow() - dt_timedelta(days=7)
    ).count()

    return {
        "usuarios": {
            "cadastrados": total_usuarios,
            "utilizando_7d": ativos
        },
        "assinaturas": {
            "em_dia": pagos,
            "vencidos": vencidos
        },
        "data": hoje.isoformat()
    }

# ======================
# ROTAS DE TEMPLATES
# ======================

# Rota para a página principal (Dashboard)
@app.get("/")
async def index(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "dashboard.html",
        get_template_context(request)
    )

@app.get("/tipos")
async def tipos(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "tipos_lancamentos.html",
        get_template_context(request)
    )

@app.get("/configuracoes")
async def configuracoes(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "configuracoes.html",
        get_template_context(request)
    )

@app.get("/metas")
async def metas_page(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "metas.html",
        get_template_context(request)
    )

@app.get("/dashboard")
async def dashboard(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "dashboard.html",
        get_template_context(request)
    )

# Rota para Lançamentos (antes era a página principal)
@app.get("/lancamentos")
async def lancamentos_page(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "lancamentos_financeiros_db.html",
        get_template_context(request)
    )

@app.get("/parcelas")
async def parcelas_page(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "parcelas_a_vencer.html",
        get_template_context(request)
    )

@app.get("/fluxo-caixa")
async def fluxo_caixa_page(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "fluxo_caixa.html",
        get_template_context(request)
    )

@app.get("/recorrentes")
async def recorrentes_page(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "recorrentes.html",
        get_template_context(request)
    )

@app.get("/formas-pagamento")
async def formas_pagamento_page(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "formas_pagamento.html",
        get_template_context(request)
    )

@app.get("/historico-pagamentos")
async def historico_pagamentos_page(request: Request, current_user: Optional[Any] = Depends(get_optional_user)):
    if not current_user:
        return RedirectResponse(url=f"/login?next={request.url.path}", status_code=307)
    return templates.TemplateResponse(
        "historico_pagamentos.html",
        get_template_context(request)
    )

# Página de Login (simples)
@app.get("/login")
async def login_page(request: Request, next: Optional[str] = "/"):
    return templates.TemplateResponse(
        "login.html",
        get_template_context(request, next=next or "/")
    )

# ======================
# SUPORTE (Contato não autenticado)
# ======================

class SupportMessageIn(BaseModel):
    nome: str = Field(..., min_length=2, max_length=100)
    assunto: str = Field(..., min_length=3, max_length=150)
    descricao: str = Field(..., min_length=5, max_length=5000)
    email: Optional[str] = Field(default=None, max_length=255)

def _send_support_email(subject: str, body: str, reply_to: Optional[str] = None) -> bool:
    host = os.getenv("SMTP_HOST")
    sender = os.getenv("SMTP_FROM")
    if not host or not sender:
        return False
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASSWORD")
    to_address = "contato@rfinance.com.br"

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = to_address
    msg["Subject"] = subject
    if reply_to:
        msg["Reply-To"] = reply_to
    msg.set_content(body)

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP(host, port, timeout=20) as server:
            try:
                server.starttls(context=context)
            except Exception:
                pass
            if user and password:
                server.login(user, password)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"Falha ao enviar email de suporte: {e}")
        return False

@app.post("/api/suporte/contato")
async def suporte_contato(msg: SupportMessageIn, request: Request):
    agora = datetime.utcnow().isoformat()
    client_ip = getattr(request.client, "host", None)
    user_agent = request.headers.get("user-agent", "")
    subject = f"[DOMO360] Suporte - {msg.assunto}"
    body = (
        f"Nome: {msg.nome}\n"
        f"Email: {msg.email or '-'}\n"
        f"IP: {client_ip or '-'}\n"
        f"User-Agent: {user_agent}\n"
        f"Data (UTC): {agora}\n\n"
        f"Descrição:\n{msg.descricao}\n"
    )

    sent = _send_support_email(subject, body, reply_to=msg.email or None)

    # Registrar em arquivo para auditoria/retentativa
    try:
        log_path = DATA_DIR / "support_messages.jsonl"
        with open(log_path, "a", encoding="utf-8") as f:
            record = {
                "timestamp": agora,
                "nome": msg.nome,
                "email": msg.email,
                "assunto": msg.assunto,
                "descricao": msg.descricao,
                "client_ip": client_ip,
                "user_agent": user_agent,
                "email_sent": sent,
            }
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except Exception as e:
        print(f"Falha ao registrar mensagem de suporte: {e}")

    if sent:
        return {"status": "ok"}
    else:
        return JSONResponse(status_code=202, content={
            "status": "accepted",
            "detail": "Mensagem registrada. Envio de e-mail não configurado.",
        })

# Página de Login Admin
@app.get("/admin/login")
async def admin_login_page(request: Request, next: Optional[str] = "/"):
    return templates.TemplateResponse(
        "admin_login.html",
        get_template_context(request, next=next or "/")
    )

@app.post("/auth/admin/login", response_model=Token)
async def admin_login(
    request: Request,
    response: Response,
    login_data: LoginRequest,
    db: Session = Depends(get_db)
):
    """
    Autentica administrador e retorna token JWT.
    Verifica se o usuário possui flag admin=True.
    Envia notificações de segurança e aplica bloqueio após 4 tentativas falhas.
    """
    from app.security_notifications import check_and_notify_login_attempt
    
    # Obter IP e User-Agent
    client_ip = request.client.host if request.client else "unknown"
    fwd = request.headers.get("x-forwarded-for") or request.headers.get("X-Forwarded-For")
    if fwd:
        client_ip = fwd.split(',')[0].strip()
    user_agent = request.headers.get("user-agent", "")
    
    # Verificar se conta está bloqueada por tentativas anteriores
    cutoff = datetime.utcnow() - timedelta(minutes=30)
    blocked_attempt = db.query(LoginAttempt).filter(
        LoginAttempt.email == login_data.email,
        LoginAttempt.is_admin_attempt == True,
        LoginAttempt.blocked_until != None,
        LoginAttempt.blocked_until > datetime.utcnow()
    ).order_by(LoginAttempt.blocked_until.desc()).first()
    
    if blocked_attempt:
        remaining = (blocked_attempt.blocked_until - datetime.utcnow()).total_seconds() / 60
        raise HTTPException(
            status_code=429,
            detail=f"Conta temporariamente bloqueada devido a múltiplas tentativas falhas. Tente novamente em {int(remaining)} minutos."
        )
    
    # Autentica usuário
    user = authenticate_user(db, login_data.email, login_data.password)
    success = user is not None and user.ativo and user.admin
    
    # Registrar tentativa e enviar notificações
    attempt_status = check_and_notify_login_attempt(
        db=db,
        email=login_data.email,
        ip=client_ip,
        user_agent=user_agent,
        success=success,
        is_admin=True
    )
    
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Credenciais inválidas"
        )
    
    if not user.ativo:
        raise HTTPException(
            status_code=403,
            detail="Usuário inativo"
        )
    
    # Verificar se é admin
    if not user.admin:
        raise HTTPException(
            status_code=403,
            detail="Acesso restrito a administradores"
        )
    
    # Se foi bloqueado nesta tentativa, informar
    if attempt_status["blocked"]:
        raise HTTPException(
            status_code=429,
            detail="Muitas tentativas falhas. Conta bloqueada temporariamente por 30 minutos."
        )
    
    # Cria token de acesso com id e email
    access_token = create_access_token(data={"sub": str(user.id), "email": user.email})
    
    # Define cookie com o token (httponly para segurança)
    cookie_params = {
        "key": "access_token",
        "value": access_token,
        "httponly": True,
        "secure": os.getenv("HTTPS_ENABLED", "false").lower() == "true",
        "samesite": "lax",
        "max_age": ACCESS_TOKEN_EXPIRE_MINUTES * 60
    }
    response.set_cookie(**cookie_params)
    
    # Atualizar último acesso
    user.ultimo_acesso = datetime.utcnow()
    db.commit()
    
    from app.auth import verify_password
    must_change = False
    # Detectar senha padrão insegura e exigir troca imediata
    try:
        if verify_password('123456', user.senha_hash):
            must_change = True
    except Exception:
        pass
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserOut.from_orm(user),
        require_password_change=must_change
    )

# ======================
# ADMIN: Páginas e APIs
# ======================

@app.get("/admin")
async def admin_home(
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user),
):
    from app.auth import verify_password
    try:
        if verify_password('123456', current_user.senha_hash):
            return RedirectResponse(url='/admin/alterar-senha', status_code=303)
    except Exception:
        pass
    return RedirectResponse(url="/admin/clientes", status_code=302)

class AdminClienteOut(BaseModel):
    id: int
    nome: str
    email: str
    primeiro_acesso: Optional[datetime]
    tipo_servico: Optional[str]
    inicio_contrato: Optional[date]
    validade_contrato: Optional[date]
    fim_contrato: Optional[date]
    inativo_dias: Optional[int]
    assinatura_status: Optional[str]

    class Config:
        from_attributes = True

class AdminClientesResponse(BaseModel):
    items: List[AdminClienteOut]
    total: int
    page: int
    page_size: int
    total_pages: int
    counts_by_status: Dict[str, int]

@app.get("/admin/clientes")
async def admin_clientes_page(
    request: Request,
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user),
):
    # Se senha padrão, redirecionar para alterar senha
    from app.auth import verify_password
    try:
        if verify_password('123456', current_user.senha_hash):
            return RedirectResponse(url='/admin/alterar-senha', status_code=303)
    except Exception:
        pass
    return templates.TemplateResponse(
        "admin_clientes.html",
        get_template_context(request)
    )

# ======================
# ADMIN: Usuários Admin (listar/criar)
# ======================

@app.get("/admin/usuarios")
async def admin_usuarios_page(
    request: Request,
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
    success: Optional[str] = None,
    error: Optional[str] = None,
):
    admins = db.query(User).filter(User.admin == True).order_by(User.created_at.desc()).all()
    return templates.TemplateResponse(
        "admin_usuarios.html",
        get_template_context(request, admins=admins, success=success, error=error)
    )

@app.post("/admin/usuarios")
async def admin_usuarios_create(
    request: Request,
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    form = await request.form()
    nome = (form.get("nome") or "").strip()
    email = (form.get("email") or "").strip().lower()
    password = (form.get("password") or "").strip()
    confirm = (form.get("confirm") or "").strip()

    if not nome or not email or not password or not confirm:
        return RedirectResponse("/admin/usuarios?error=Preencha todos os campos.", status_code=303)
    if password != confirm:
        return RedirectResponse("/admin/usuarios?error=Confirmação de senha não confere.", status_code=303)
    # Reforço: senha forte mínima 12 + complexidade
    from app.auth import validate_password_strength
    if len(password) < 12 or not validate_password_strength(password):
        return RedirectResponse("/admin/usuarios?error=Senha fraca: mínimo 12 caracteres, com maiúscula, minúscula, dígito e símbolo.", status_code=303)

    try:
        from app.auth import UserCreate, create_user
        created = create_user(db, UserCreate(nome=nome, email=email, password=password), is_admin=True)
        return RedirectResponse(f"/admin/usuarios?success=Admin criado: {created.email}", status_code=303)
    except ValueError as ve:
        # Mensagem genérica para não vazar existência de email
        return RedirectResponse("/admin/usuarios?error=Não foi possível criar o admin. Verifique os dados.", status_code=303)
    except Exception:
        return RedirectResponse("/admin/usuarios?error=Falha inesperada ao criar admin.", status_code=303)

# ======================
# ADMIN: Teste de Alertas
# ======================

class TestAlertRequest(BaseModel):
    channel: str = Field(..., pattern="^(sms|whatsapp|email)$")
    message: Optional[str] = None

@app.post("/api/admin/test-alert")
async def api_admin_test_alert(
    req: TestAlertRequest,
    request: Request,
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user),
):
    client_ip = request.headers.get("x-forwarded-for") or (request.client.host if request.client else "unknown")
    user_agent = request.headers.get("user-agent", "")
    target_email = os.getenv("ADMIN_ALERT_EMAIL", current_user.email)
    target_phone = os.getenv("ADMIN_ALERT_PHONE")
    target_whatsapp = os.getenv("ADMIN_ALERT_WHATSAPP")
    message = req.message or "Teste de alerta de segurança (DOMO360)"

    results: Dict[str, Any] = {}
    try:
        if req.channel == "sms":
            if not target_phone:
                return JSONResponse(status_code=400, content={"detail": "ADMIN_ALERT_PHONE não configurado"})
            ok = send_login_alert_sms(target_phone, current_user.email, client_ip, success=False)
            results["sms"] = ok
        elif req.channel == "whatsapp":
            if not target_whatsapp:
                return JSONResponse(status_code=400, content={"detail": "ADMIN_ALERT_WHATSAPP não configurado"})
            ok = send_login_alert_whatsapp(target_whatsapp, current_user.email, client_ip, success=False)
            results["whatsapp"] = ok
        else:
            ok = send_login_alert_email(target_email, current_user.email, client_ip, user_agent, success=False)
            results["email"] = ok
        return {"ok": True, "results": results}
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": str(e)})

@app.get("/admin/alterar-senha")
async def admin_alterar_senha_page(
    request: Request,
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user),
    success: Optional[str] = None,
    error: Optional[str] = None,
):
    return templates.TemplateResponse(
        "admin_alterar_senha.html",
        get_template_context(request, success=success, error=error)
    )

## Endpoint de emergência removido após uso (segurança restaurada)

@app.post("/admin/alterar-senha")
async def admin_alterar_senha_submit(
    request: Request,
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    form = await request.form()
    senha_atual = (form.get("senha_atual") or "").strip()
    nova_senha = (form.get("nova_senha") or "").strip()
    confirmar = (form.get("confirmar_senha") or "").strip()

    if not senha_atual or not nova_senha or not confirmar:
        return templates.TemplateResponse(
            "admin_alterar_senha.html",
            get_template_context(request, error="Preencha todos os campos."),
            status_code=400
        )
    if nova_senha != confirmar:
        return templates.TemplateResponse(
            "admin_alterar_senha.html",
            get_template_context(request, error="Confirmação não confere."),
            status_code=400
        )
    try:
        from app.auth import change_admin_password
        change_admin_password(db, current_user.id, senha_atual, nova_senha)
        # Opcional: invalidar sessão atual forçando novo login? Mantemos para comodidade.
        return RedirectResponse(url="/admin/alterar-senha?success=1", status_code=303)
    except ValueError as ve:
        return templates.TemplateResponse(
            "admin_alterar_senha.html",
            get_template_context(request, error=str(ve)),
            status_code=400
        )
    except Exception:
        return templates.TemplateResponse(
            "admin_alterar_senha.html",
            get_template_context(request, error="Falha inesperada."),
            status_code=500
        )

    # ======================
    # ADMIN: Administração de usuários (consulta)
    # ======================

    @app.get("/api/admin/admins/count")
    async def api_admin_admins_count(
        _ip_ok: bool = Depends(ensure_admin_ip_allowed),
        current_user: User = Depends(get_current_admin_user),
        db: Session = Depends(get_db),
    ):
        total = db.query(User).filter(User.admin == True).count()
        return {"count": total}

    @app.get("/api/admin/admins")
    async def api_admin_admins_list(
        _ip_ok: bool = Depends(ensure_admin_ip_allowed),
        current_user: User = Depends(get_current_admin_user),
        db: Session = Depends(get_db),
    ):
        admins = db.query(User).filter(User.admin == True).all()
        emails = [u.email for u in admins]
        return {"count": len(emails), "emails": emails}

@app.get("/api/admin/clientes", response_model=AdminClientesResponse)
async def api_admin_clientes(
    only_active: bool = True,
    q: Optional[str] = None,
    status: Optional[str] = None,  # trial | ativa | inadimplente | cancelada | sem
    order_by: str = "nome",
    order_dir: str = "asc",  # asc | desc
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _ip_ok: bool = Depends(ensure_admin_ip_allowed),
    current_user: User = Depends(get_current_admin_user)
):
    # Bounds
    if page < 1:
        page = 1
    page_size = max(1, min(100, page_size))

    users = list_users(db, skip=0, limit=5000, only_active=only_active)

    # Carregar assinaturas em lote
    user_ids = [u.id for u in users]
    assinaturas_map: Dict[int, Assinatura] = {}
    if user_ids:
        subs = db.query(Assinatura).filter(Assinatura.usuario_id.in_(user_ids)).all()
        assinaturas_map = {s.usuario_id: s for s in subs}

    out: List[AdminClienteOut] = []
    today = datetime.utcnow().date()
    for u in users:
        sub = assinaturas_map.get(u.id)
        ultimo = u.ultimo_acesso.date() if u.ultimo_acesso else None
        base_inativo = ultimo or (u.created_at.date() if u.created_at else None)
        inativo_dias = (today - base_inativo).days if base_inativo else None

        out.append(AdminClienteOut(
            id=u.id,
            nome=u.nome,
            email=u.email,
            primeiro_acesso=u.created_at,
            tipo_servico=None,
            inicio_contrato=sub.data_inicio if sub else None,
            validade_contrato=sub.proximo_vencimento if sub else None,
            fim_contrato=sub.cancelada_em if (sub and sub.cancelada_em) else None,
            inativo_dias=inativo_dias,
            assinatura_status=sub.status if sub else None,
        ))

    # Filtro por busca
    if q:
        ql = q.lower().strip()
        out = [c for c in out if (c.nome or '').lower().find(ql) >= 0 or (c.email or '').lower().find(ql) >= 0]

    # Filtro por status de assinatura
    if status:
        st = status.lower()
        if st == "sem":
            out = [c for c in out if not c.assinatura_status]
        else:
            out = [c for c in out if (c.assinatura_status or '').lower() == st]

    # Contagem por status (antes de paginar)
    counts: Dict[str, int] = {"trial":0, "ativa":0, "inadimplente":0, "cancelada":0, "sem":0}
    for c in out:
        key = (c.assinatura_status or 'sem').lower()
        if key not in counts:
            counts[key] = 0
        counts[key] += 1

    # Ordenação
    valid_fields = {
        "nome": lambda c: (c.nome or "").lower(),
        "email": lambda c: (c.email or "").lower(),
        "primeiro_acesso": lambda c: c.primeiro_acesso or datetime.min,
        "inicio_contrato": lambda c: c.inicio_contrato or date.min,
        "validade_contrato": lambda c: c.validade_contrato or date.min,
        "fim_contrato": lambda c: c.fim_contrato or date.min,
        "inativo_dias": lambda c: c.inativo_dias if c.inativo_dias is not None else 10**9,
        "assinatura_status": lambda c: (c.assinatura_status or "zzzz")
    }
    key_fn = valid_fields.get(order_by, valid_fields["nome"])
    reverse = (order_dir.lower() == "desc")
    out.sort(key=key_fn, reverse=reverse)

    # Paginação
    total = len(out)
    total_pages = (total + page_size - 1) // page_size if total else 1
    start = (page - 1) * page_size
    end = start + page_size
    items = out[start:end]

    return AdminClientesResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        counts_by_status=counts,
    )

# Página de Registro
@app.get("/register")
async def register_page(request: Request):
    return templates.TemplateResponse(
        "register.html",
        get_template_context(request)
    )

# ======================
# ENDPOINTS DE FORMAS DE PAGAMENTO
# ======================

@app.get("/api/formas-pagamento", response_model=List[FormaPagamentoOut])
async def listar_formas_pagamento(
    incluir_inativas: bool = False,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Lista todas as formas de pagamento do usuário autenticado"""
    query = db.query(FormaPagamento).filter(FormaPagamento.usuario_id == current_user.id)
    
    if not incluir_inativas:
        query = query.filter(FormaPagamento.ativo == True)
    
    formas = query.order_by(FormaPagamento.nome).all()
    return [FormaPagamentoOut.from_orm(f) for f in formas]

@app.get("/api/formas-pagamento/{forma_id}", response_model=FormaPagamentoOut)
async def obter_forma_pagamento(
    forma_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Obtém uma forma de pagamento específica do usuário autenticado"""
    forma = db.query(FormaPagamento).filter(
        FormaPagamento.id == forma_id,
        FormaPagamento.usuario_id == current_user.id
    ).first()
    if not forma:
        raise HTTPException(status_code=404, detail="Forma de pagamento não encontrada")
    return FormaPagamentoOut.from_orm(forma)

@app.post("/api/formas-pagamento", response_model=FormaPagamentoOut, status_code=201)
async def criar_forma_pagamento(
    forma: FormaPagamentoIn,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """Cria uma nova forma de pagamento para o usuário autenticado"""
    from datetime import date
    
    # Verificar se já existe uma forma com o mesmo nome para este usuário
    forma_existente = db.query(FormaPagamento).filter(
        FormaPagamento.nome == forma.nome,
        FormaPagamento.usuario_id == current_user.id
    ).first()
    
    if forma_existente:
        raise HTTPException(
            status_code=400,
            detail=f"Já existe uma forma de pagamento com o nome '{forma.nome}'"
        )
    
    try:
        nova_forma = FormaPagamento(
            usuario_id=current_user.id,
            nome=forma.nome,
            tipo=forma.tipo,
            banco=forma.banco,
            limite_credito="{:.2f}".format(forma.limite_credito) if forma.limite_credito else None,
            ativo=forma.ativo,
            created_at=date.today(),
            observacao=forma.observacao
        )
        
        db.add(nova_forma)
        db.commit()
        db.refresh(nova_forma)
        
        return FormaPagamentoOut.from_orm(nova_forma)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao criar forma de pagamento: {str(e)}")

@app.put("/api/formas-pagamento/{forma_id}", response_model=FormaPagamentoOut)
async def atualizar_forma_pagamento(forma_id: int, forma: FormaPagamentoIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Atualiza uma forma de pagamento existente"""
    db_forma = db.query(FormaPagamento).filter(
        FormaPagamento.id == forma_id,
        FormaPagamento.usuario_id == current_user.id
    ).first()
    
    if not db_forma:
        raise HTTPException(status_code=404, detail="Forma de pagamento não encontrada")
    
    # Verificar se o novo nome já existe (exceto para o próprio registro)
    if forma.nome != db_forma.nome:
        forma_existente = db.query(FormaPagamento).filter(
            FormaPagamento.nome == forma.nome,
            FormaPagamento.id != forma_id,
            FormaPagamento.usuario_id == current_user.id
        ).first()
        
        if forma_existente:
            raise HTTPException(
                status_code=400,
                detail=f"Já existe uma forma de pagamento com o nome '{forma.nome}'"
            )
    
    try:
        db_forma.nome = forma.nome
        db_forma.tipo = forma.tipo
        db_forma.banco = forma.banco
        db_forma.limite_credito = "{:.2f}".format(forma.limite_credito) if forma.limite_credito else None
        db_forma.ativo = forma.ativo
        db_forma.observacao = forma.observacao
        
        db.commit()
        db.refresh(db_forma)
        
        return FormaPagamentoOut.from_orm(db_forma)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar forma de pagamento: {str(e)}")

@app.delete("/api/formas-pagamento/{forma_id}")
async def excluir_forma_pagamento(forma_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Exclui uma forma de pagamento (se não estiver em uso)"""
    # Verifica se a forma está sendo usada em alguma parcela paga
    parcelas_com_forma = db.query(Parcela).filter(Parcela.forma_pagamento_id == forma_id).count()
    
    if parcelas_com_forma > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Não é possível excluir esta forma de pagamento pois ela está sendo usada em {parcelas_com_forma} parcela(s)"
        )
    
    forma = db.query(FormaPagamento).filter(
        FormaPagamento.id == forma_id,
        FormaPagamento.usuario_id == current_user.id
    ).first()
    if not forma:
        raise HTTPException(status_code=404, detail="Forma de pagamento não encontrada")
    
    try:
        db.delete(forma)
        db.commit()
        return {"status": "ok", "message": "Forma de pagamento excluída com sucesso"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao excluir forma de pagamento: {str(e)}")

@app.patch("/api/formas-pagamento/{forma_id}/toggle")
async def toggle_forma_pagamento(forma_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Ativa/desativa uma forma de pagamento"""
    forma = db.query(FormaPagamento).filter(
        FormaPagamento.id == forma_id,
        FormaPagamento.usuario_id == current_user.id
    ).first()
    
    if not forma:
        raise HTTPException(status_code=404, detail="Forma de pagamento não encontrada")
    
    try:
        forma.ativo = not forma.ativo
        db.commit()
        db.refresh(forma)
        
        return FormaPagamentoOut.from_orm(forma)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao alternar status: {str(e)}")

@app.get("/api/formas-pagamento/{forma_id}/usage")
async def obter_uso_forma_pagamento(forma_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Retorna quantas parcelas estão usando esta forma de pagamento"""
    forma = db.query(FormaPagamento).filter(
        FormaPagamento.id == forma_id,
        FormaPagamento.usuario_id == current_user.id
    ).first()
    if not forma:
        raise HTTPException(status_code=404, detail="Forma de pagamento não encontrada")
    
    # Contar parcelas que usam esta forma
    count = db.query(Parcela).filter(
        Parcela.forma_pagamento_id == forma_id,
        Parcela.usuario_id == current_user.id
    ).count()
    
    return {
        "forma_id": forma_id,
        "forma_nome": forma.nome,
        "parcelas_vinculadas": count,
        "em_uso": count > 0
    }

@app.get("/api/tipos", response_model=List[TipoLancamentoOut])
async def listar_tipos(current_user: User = Depends(get_current_active_user), db: Session = Depends(get_db)):
    try:
        tipos = (
            db.query(TipoLancamento)
            .filter(TipoLancamento.usuario_id == current_user.id)
            .order_by(TipoLancamento.nome)
            .all()
        )
        
        result = []
        for tipo in tipos:
            try:
                tipo_out = TipoLancamentoOut.model_validate(tipo)
                result.append(tipo_out)
            except Exception as e:
                print(f"Erro ao converter tipo {tipo.id}: {str(e)}")
                raise
                
        return result
    except Exception as e:
        print(f"Erro geral: {str(e)}")
        raise

@app.post("/api/tipos", response_model=TipoLancamentoOut)
async def criar_tipo(tipo: TipoLancamentoIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    from datetime import date, timedelta
    
    # Verificar se já existe um tipo com o mesmo nome e natureza para este usuário
    tipo_existente = db.query(TipoLancamento).filter(
        TipoLancamento.usuario_id == current_user.id,
        TipoLancamento.nome == tipo.nome,
        TipoLancamento.natureza == tipo.natureza
    ).first()
    
    if tipo_existente:
        raise HTTPException(
            status_code=400,
            detail=f"Já existe um tipo '{tipo.nome}' de natureza '{tipo.natureza}'"
        )
    
    db_tipo = TipoLancamento(
        usuario_id=current_user.id,
        nome=tipo.nome,
        natureza=tipo.natureza,
        created_at=date.today()
    )
    db.add(db_tipo)
    db.commit()
    db.refresh(db_tipo)
    return TipoLancamentoOut.from_orm(db_tipo)

@app.delete("/api/tipos/{tipo_id}")
async def excluir_tipo(tipo_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    # Verifica se o tipo está sendo usado em algum lançamento
    if db.query(Lancamento).filter(Lancamento.tipo_lancamento_id == tipo_id).first():
        raise HTTPException(
            status_code=400,
            detail="Não é possível excluir este tipo pois ele está sendo usado em lançamentos"
        )
    
    tipo = db.query(TipoLancamento).filter(
        TipoLancamento.id == tipo_id,
        TipoLancamento.usuario_id == current_user.id
    ).first()
    if not tipo:
        raise HTTPException(status_code=404, detail="Tipo não encontrado")
    
    db.delete(tipo)
    db.commit()
    return {"status": "ok"}

# ======================
# ENDPOINTS DE SUBTIPOS
# ======================

@app.get("/api/tipos/{tipo_id}/subtipos", response_model=List[SubtipoLancamentoOut])
async def listar_subtipos(tipo_id: int, current_user: User = Depends(get_current_active_user), db: Session = Depends(get_db)):
    """Lista todos os subtipos de um tipo específico"""
    # Verifica se o tipo existe
    tipo = db.query(TipoLancamento).filter(
        TipoLancamento.id == tipo_id,
        TipoLancamento.usuario_id == current_user.id
    ).first()
    if not tipo:
        raise HTTPException(status_code=404, detail="Tipo não encontrado")
    
    subtipos = db.query(SubtipoLancamento).filter(
        SubtipoLancamento.usuario_id == current_user.id,
        SubtipoLancamento.tipo_lancamento_id == tipo_id
    ).order_by(SubtipoLancamento.nome).all()
    
    return [SubtipoLancamentoOut.model_validate(s) for s in subtipos]

@app.get("/api/subtipos", response_model=List[SubtipoLancamentoOut])
async def listar_todos_subtipos(current_user: User = Depends(get_current_active_user), db: Session = Depends(get_db)):
    """Lista todos os subtipos (de todos os tipos)"""
    subtipos = (
        db.query(SubtipoLancamento)
        .filter(SubtipoLancamento.usuario_id == current_user.id)
        .order_by(SubtipoLancamento.tipo_lancamento_id, SubtipoLancamento.nome)
        .all()
    )
    
    return [SubtipoLancamentoOut.model_validate(s) for s in subtipos]

@app.post("/api/tipos/{tipo_id}/subtipos", response_model=SubtipoLancamentoOut, status_code=201)
async def criar_subtipo(tipo_id: int, subtipo: SubtipoLancamentoIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Cria um novo subtipo para um tipo específico"""
    from datetime import date
    
    # Verifica se o tipo existe e pertence ao usuário
    tipo = db.query(TipoLancamento).filter(
        TipoLancamento.id == tipo_id,
        TipoLancamento.usuario_id == current_user.id
    ).first()
    if not tipo:
        raise HTTPException(status_code=404, detail="Tipo não encontrado")
    
    # Verifica se tipo_lancamento_id do body corresponde ao da URL
    if subtipo.tipo_lancamento_id != tipo_id:
        raise HTTPException(
            status_code=400,
            detail="tipo_lancamento_id não corresponde ao ID da URL"
        )
    
    # Verifica se já existe subtipo com mesmo nome para este tipo
    subtipo_existente = db.query(SubtipoLancamento).filter(
        SubtipoLancamento.usuario_id == current_user.id,
        SubtipoLancamento.tipo_lancamento_id == tipo_id,
        SubtipoLancamento.nome == subtipo.nome
    ).first()
    
    if subtipo_existente:
        raise HTTPException(
            status_code=400,
            detail=f"Já existe um subtipo '{subtipo.nome}' para este tipo"
        )
    
    db_subtipo = SubtipoLancamento(
        usuario_id=current_user.id,
        tipo_lancamento_id=tipo_id,
        nome=subtipo.nome,
        ativo=subtipo.ativo,
        created_at=date.today()
    )
    db.add(db_subtipo)
    db.commit()
    db.refresh(db_subtipo)
    
    return SubtipoLancamentoOut.model_validate(db_subtipo)

@app.patch("/api/subtipos/{subtipo_id}", response_model=SubtipoLancamentoOut)
async def atualizar_subtipo(subtipo_id: int, subtipo_update: SubtipoLancamentoIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Atualiza um subtipo existente"""
    db_subtipo = db.query(SubtipoLancamento).filter(
        SubtipoLancamento.id == subtipo_id,
        SubtipoLancamento.usuario_id == current_user.id
    ).first()
    
    if not db_subtipo:
        raise HTTPException(status_code=404, detail="Subtipo não encontrado")
    
    # Verifica se o novo nome já existe para o mesmo tipo (exceto o próprio registro)
    if subtipo_update.nome != db_subtipo.nome:
        subtipo_existente = db.query(SubtipoLancamento).filter(
            SubtipoLancamento.usuario_id == current_user.id,
            SubtipoLancamento.tipo_lancamento_id == subtipo_update.tipo_lancamento_id,
            SubtipoLancamento.nome == subtipo_update.nome,
            SubtipoLancamento.id != subtipo_id
        ).first()
        
        if subtipo_existente:
            raise HTTPException(
                status_code=400,
                detail=f"Já existe um subtipo '{subtipo_update.nome}' para este tipo"
            )
    
    db_subtipo.nome = subtipo_update.nome
    db_subtipo.ativo = subtipo_update.ativo
    db_subtipo.tipo_lancamento_id = subtipo_update.tipo_lancamento_id
    
    db.commit()
    db.refresh(db_subtipo)
    
    return SubtipoLancamentoOut.model_validate(db_subtipo)

@app.delete("/api/subtipos/{subtipo_id}")
async def excluir_subtipo(subtipo_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Exclui um subtipo (se não estiver em uso)"""
    # Verifica se o subtipo está sendo usado em algum lançamento
    if db.query(Lancamento).filter(Lancamento.subtipo_lancamento_id == subtipo_id).first():
        raise HTTPException(
            status_code=400,
            detail="Não é possível excluir este subtipo pois ele está sendo usado em lançamentos"
        )
    
    subtipo = db.query(SubtipoLancamento).filter(
        SubtipoLancamento.id == subtipo_id,
        SubtipoLancamento.usuario_id == current_user.id
    ).first()
    if not subtipo:
        raise HTTPException(status_code=404, detail="Subtipo não encontrado")
    
    db.delete(subtipo)
    db.commit()
    return {"status": "ok"}

# Rotas da API
@app.post("/api/lancamentos")
async def criar_lancamento(
    lancamento: LancamentoIn,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    from datetime import date
    import traceback
    
    try:
        print(f"\n=== Recebendo lançamento ===")
        print(f"Dados recebidos: {lancamento.model_dump()}")
        print(f"Usuário: {current_user.email} (ID: {current_user.id})")
        
        # Validar o tipo_lancamento_id se fornecido (e se pertence ao usuário)
        if lancamento.tipo_lancamento_id:
            print(f"Validando tipo_lancamento_id: {lancamento.tipo_lancamento_id}")
            tipo = db.query(TipoLancamento).filter(
                TipoLancamento.id == lancamento.tipo_lancamento_id,
                TipoLancamento.usuario_id == current_user.id
            ).first()
            if not tipo:
                print(f"Tipo de lançamento {lancamento.tipo_lancamento_id} não encontrado")
                raise HTTPException(status_code=404, detail=f"Tipo de lançamento ID {lancamento.tipo_lancamento_id} não encontrado")
            if tipo.natureza != lancamento.tipo:
                print(f"Natureza incompatível: tipo.natureza={tipo.natureza}, lancamento.tipo={lancamento.tipo}")
                raise HTTPException(status_code=400, detail=f"O tipo de lançamento selecionado é '{tipo.natureza}' mas você está criando uma '{lancamento.tipo}'")
            print(f"Tipo validado: {tipo.nome} ({tipo.natureza})")
        
        # Validar o subtipo_lancamento_id se fornecido
        if lancamento.subtipo_lancamento_id:
            print(f"Validando subtipo_lancamento_id: {lancamento.subtipo_lancamento_id}")
            subtipo = db.query(SubtipoLancamento).filter(SubtipoLancamento.id == lancamento.subtipo_lancamento_id).first()
            if not subtipo:
                print(f"Subtipo de lançamento {lancamento.subtipo_lancamento_id} não encontrado")
                raise HTTPException(status_code=404, detail=f"Subtipo de lançamento ID {lancamento.subtipo_lancamento_id} não encontrado")
            if not subtipo.ativo:
                raise HTTPException(status_code=400, detail=f"O subtipo '{subtipo.nome}' está inativo")
            # Validar se o subtipo pertence ao tipo selecionado (se tipo foi informado)
            if lancamento.tipo_lancamento_id and subtipo.tipo_lancamento_id != lancamento.tipo_lancamento_id:
                raise HTTPException(status_code=400, detail=f"O subtipo '{subtipo.nome}' não pertence ao tipo selecionado")
            print(f"Subtipo validado: {subtipo.nome}")

        print(f"Valor total: {lancamento.valor_total} (tipo: {type(lancamento.valor_total).__name__})")
        print(f"Valor médio: {lancamento.valor_medio_parcelas} (tipo: {type(lancamento.valor_medio_parcelas).__name__})")
        
        # Converter valores numéricos para string com 2 casas decimais para o SQLAlchemy
        valor_total_str = "{:.2f}".format(lancamento.valor_total)
        valor_medio_str = "{:.2f}".format(lancamento.valor_medio_parcelas)
        
        print(f"Criando objeto Lancamento no banco...")
        db_lancamento = Lancamento(
            usuario_id=current_user.id,
            data_lancamento=date.fromisoformat(lancamento.data_lancamento),
            tipo=lancamento.tipo,
            tipo_lancamento_id=lancamento.tipo_lancamento_id,
            subtipo_lancamento_id=lancamento.subtipo_lancamento_id,
            fornecedor=lancamento.fornecedor,
            valor_total=valor_total_str,
            data_primeiro_vencimento=date.fromisoformat(lancamento.data_primeiro_vencimento),
            numero_parcelas=lancamento.numero_parcelas,
            valor_medio_parcelas=valor_medio_str,
            observacao=lancamento.observacao
        )
        
        try:
            db.add(db_lancamento)
            # Use flush to obtain the ID without committing; commit only after creating parcels (atomic operation)
            db.flush()
            db.refresh(db_lancamento)
            print(f"✓ Lançamento preparado (ID obtido) para criar parcelas. ID: {db_lancamento.id}")
            
            # Gerar parcelas automaticamente
            from dateutil.relativedelta import relativedelta
            from decimal import Decimal
            data_venc = date.fromisoformat(lancamento.data_primeiro_vencimento)
            parcelas_criadas = []
            
            # Calcular valores com precisão
            valor_total_decimal = Decimal(str(lancamento.valor_total))
            valor_parcela_base = valor_total_decimal / lancamento.numero_parcelas
            valor_parcela_arredondado = valor_parcela_base.quantize(Decimal('0.01'))
            
            # Calcular diferença (centavos restantes)
            soma_parcelas = valor_parcela_arredondado * lancamento.numero_parcelas
            diferenca = valor_total_decimal - soma_parcelas
            centavos_restantes = int(diferenca * 100)  # Converter para centavos
            
            print(f"Valor total: {valor_total_decimal}, Parcela base: {valor_parcela_arredondado}, Diferença: {diferenca} ({centavos_restantes} centavos)")
            
            for i in range(lancamento.numero_parcelas):
                valor_parcela = valor_parcela_arredondado
                
                # Distribuir centavos restantes nas últimas parcelas
                if centavos_restantes != 0:
                    parcelas_faltantes = lancamento.numero_parcelas - i
                    if parcelas_faltantes <= abs(centavos_restantes):
                        ajuste = Decimal('0.01') if centavos_restantes > 0 else Decimal('-0.01')
                        valor_parcela += ajuste
                        centavos_restantes -= (1 if centavos_restantes > 0 else -1)
                
                parcela = Parcela(
                    usuario_id=current_user.id,
                    lancamento_id=db_lancamento.id,
                    numero_parcela=i + 1,
                    data_vencimento=data_venc + relativedelta(months=i),
                    valor="{:.2f}".format(valor_parcela),
                    paga=0,
                    data_pagamento=None
                )
                db.add(parcela)
                parcelas_criadas.append(parcela)
            
            db.commit()
            print(f"✓ {len(parcelas_criadas)} parcelas criadas!")
            print(f"=== Fim da criação ===\n")
            
            # Adicionar parcelas ao objeto para retorno
            db_lancamento._parcelas = parcelas_criadas
            # Retornar como dict para evitar qualquer incompatibilidade de serialização
            return LancamentoOut.from_orm(db_lancamento, incluir_parcelas=True).model_dump()
        except Exception as db_error:
            db.rollback()
            error_msg = str(db_error)
            print(f"✗ Erro ao salvar no banco: {error_msg}")
            print(f"Traceback: {traceback.format_exc()}")
            raise HTTPException(status_code=500, detail=f"Erro ao salvar no banco de dados: {error_msg}")
            
    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e)
        print(f"✗ Erro geral ao processar lançamento: {error_msg}")
        print(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar lançamento: {error_msg}")

@app.get("/api/lancamentos", response_model=List[LancamentoOut])
async def listar_lancamentos(
    tipo: Optional[str] = None,
    tipo_lancamento_id: Optional[int] = None,
    subtipo_lancamento_id: Optional[int] = None,
    fornecedor: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    try:
        print(f"Listando lançamentos do usuário {current_user.id}...")
        query = db.query(Lancamento).filter(Lancamento.usuario_id == current_user.id)
        
        # Aplicar filtros
        if tipo:
            query = query.filter(Lancamento.tipo == tipo)
        if tipo_lancamento_id:
            query = query.filter(Lancamento.tipo_lancamento_id == tipo_lancamento_id)
        if subtipo_lancamento_id:
            query = query.filter(Lancamento.subtipo_lancamento_id == subtipo_lancamento_id)
        if fornecedor:
            query = query.filter(Lancamento.fornecedor.ilike(f"%{fornecedor}%"))
        if data_inicio:
            from datetime import date
            query = query.filter(Lancamento.data_lancamento >= date.fromisoformat(data_inicio))
        if data_fim:
            from datetime import date
            query = query.filter(Lancamento.data_lancamento <= date.fromisoformat(data_fim))
        
        lancamentos = []
        for l in query.order_by(Lancamento.id.desc()).all():
            try:
                lancamento_out = LancamentoOut.from_orm(l)
                lancamentos.append(lancamento_out)
            except Exception as e:
                print(f"Erro ao processar lançamento {l.id}: {str(e)}")
                print(f"Traceback: {traceback.format_exc()}")
                raise
        print(f"Total de lançamentos: {len(lancamentos)}")
        return lancamentos
    except Exception as e:
        print(f"Erro ao listar lançamentos: {str(e)}")
        raise

@app.get("/api/lancamentos/{lancamento_id}", response_model=LancamentoOut)
async def obter_lancamento(lancamento_id: int, incluir_parcelas: bool = False, current_user: User = Depends(get_current_active_user), db: Session = Depends(get_db)):
    lancamento = db.query(Lancamento).filter(
        Lancamento.id == lancamento_id,
        Lancamento.usuario_id == current_user.id
    ).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    
    if incluir_parcelas:
        parcelas = db.query(Parcela).filter(Parcela.lancamento_id == lancamento_id).order_by(Parcela.numero_parcela).all()
        lancamento._parcelas = parcelas
        return LancamentoOut.from_orm(lancamento, incluir_parcelas=True)
    
    return LancamentoOut.from_orm(lancamento)

@app.get("/api/lancamentos/{lancamento_id}/parcelas", response_model=List[ParcelaOut])
async def listar_parcelas(lancamento_id: int, current_user: User = Depends(get_current_active_user), db: Session = Depends(get_db)):
    # Verificar se o lançamento existe
    lancamento = db.query(Lancamento).filter(
        Lancamento.id == lancamento_id,
        Lancamento.usuario_id == current_user.id
    ).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    
    parcelas = db.query(Parcela).filter(
        Parcela.lancamento_id == lancamento_id,
        Parcela.usuario_id == current_user.id
    ).order_by(Parcela.numero_parcela).all()
    return [ParcelaOut.from_orm(p) for p in parcelas]

@app.get("/api/parcelas/a-vencer")
async def parcelas_a_vencer(
    data_inicio: str,
    data_fim: str,
    tipo: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    from datetime import date as dt_date, timedelta
    
    # Converter datas
    data_inicio_obj = dt_date.fromisoformat(data_inicio)
    data_fim_obj = dt_date.fromisoformat(data_fim)
    hoje = dt_date.today()

    # Query base: parcelas não pagas do usuário
    from sqlalchemy import func
    query = db.query(
        Parcela.id,
        Parcela.lancamento_id,
        Parcela.numero_parcela,
        Parcela.data_vencimento,
        Parcela.valor,
        Lancamento.tipo,
        Lancamento.fornecedor,
        Lancamento.tipo_lancamento_id,
        Lancamento.subtipo_lancamento_id,
        func.coalesce(TipoLancamento.nome, 'Sem tipo').label('tipo_nome'),
        func.coalesce(SubtipoLancamento.nome, 'Sem subtipo').label('subtipo_nome')
    ).join(
        Lancamento, Parcela.lancamento_id == Lancamento.id
    ).outerjoin(
        TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
    ).outerjoin(
        SubtipoLancamento, Lancamento.subtipo_lancamento_id == SubtipoLancamento.id
    ).filter(
        Parcela.paga == 0,
        Parcela.usuario_id == current_user.id
    )

    # Filtro por tipo
    if tipo:
        query = query.filter(Lancamento.tipo == tipo)

    # Regras de data conforme status
    if status == "vencidas":
        # Mostrar todas as vencidas até hoje (ou até data_fim, se menor). Ignora limite inferior
        limite_superior = data_fim_obj if data_fim_obj <= hoje else hoje - timedelta(days=0)
        query = query.filter(Parcela.data_vencimento <= limite_superior)
    elif status == "vence_hoje":
        query = query.filter(Parcela.data_vencimento == hoje)
    elif status == "a_vencer":
        # Apenas futuras a partir de amanhã ou do data_inicio, o que for maior
        inicio_avencer = data_inicio_obj if data_inicio_obj > hoje else hoje + timedelta(days=1)
        query = query.filter(
            Parcela.data_vencimento >= inicio_avencer,
            Parcela.data_vencimento <= data_fim_obj
        )
    else:
        # Sem status (Todos): todas vencidas + hoje + próximos 30 dias
        limite_futuro = hoje + timedelta(days=30)
        query = query.filter(Parcela.data_vencimento <= limite_futuro)

    # Ordenar por data de vencimento
    results = query.order_by(Parcela.data_vencimento, Parcela.lancamento_id).all()
    
    # Formatar resultados
    parcelas = []
    for r in results:
        parcelas.append({
            "id": r.id,
            "lancamento_id": r.lancamento_id,
            "numero_parcela": r.numero_parcela,
            "data_vencimento": r.data_vencimento.isoformat(),
            "valor": float(r.valor),
            "tipo": r.tipo,
            "fornecedor": r.fornecedor,
            "tipo_lancamento_id": r.tipo_lancamento_id,
            "subtipo_lancamento_id": r.subtipo_lancamento_id,
            "tipo_nome": r.tipo_nome,
            "subtipo_nome": r.subtipo_nome
        })
    
    # Calcular estatísticas
    stats = {
        "total": len(parcelas),
        "receitas": sum(1 for p in parcelas if p["tipo"] == "receita"),
        "despesas": sum(1 for p in parcelas if p["tipo"] == "despesa"),
        "vencidas": sum(1 for p in parcelas if p["data_vencimento"] < hoje.isoformat()),
        "vence_hoje": sum(1 for p in parcelas if p["data_vencimento"] == hoje.isoformat()),
        "a_vencer": sum(1 for p in parcelas if p["data_vencimento"] > hoje.isoformat()),
        "valor_vencidas": sum(p["valor"] for p in parcelas if p["data_vencimento"] < hoje.isoformat()),
        "valor_vence_hoje": sum(p["valor"] for p in parcelas if p["data_vencimento"] == hoje.isoformat()),
        "valor_a_vencer": sum(p["valor"] for p in parcelas if p["data_vencimento"] > hoje.isoformat()),
        "valor_receitas_a_vencer": sum(p["valor"] for p in parcelas if p["data_vencimento"] > hoje.isoformat() and p["tipo"] == "receita"),
        "valor_despesas_a_vencer": sum(p["valor"] for p in parcelas if p["data_vencimento"] > hoje.isoformat() and p["tipo"] == "despesa")
    }
    
    return {
        "parcelas": parcelas,
        "stats": stats
    }

@app.get("/api/parcelas/pagas")
async def parcelas_pagas(
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    tipo: Optional[str] = None,
    forma_pagamento_id: Optional[int] = None,
    valor_min: Optional[float] = None,
    valor_max: Optional[float] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Retorna histórico de parcelas pagas do usuário autenticado
    """
    from datetime import date as dt_date, timedelta
    
    # Query base: parcelas pagas do usuário
    query = db.query(
        Parcela.id,
        Parcela.lancamento_id,
        Parcela.numero_parcela,
        Parcela.data_vencimento,
        Parcela.data_pagamento,
        Parcela.valor,
        Parcela.valor_pago,
        Parcela.forma_pagamento_id,
        Parcela.observacao_pagamento,
        Lancamento.tipo,
        Lancamento.fornecedor,
        FormaPagamento.nome.label('forma_pagamento_nome'),
        FormaPagamento.tipo.label('forma_pagamento_tipo')
    ).join(
        Lancamento, Parcela.lancamento_id == Lancamento.id
    ).outerjoin(
        FormaPagamento, Parcela.forma_pagamento_id == FormaPagamento.id
    ).filter(
        Parcela.paga == 1,
        Parcela.usuario_id == current_user.id
    )
    
    # Filtros
    if tipo:
        query = query.filter(Lancamento.tipo == tipo)
    
    if forma_pagamento_id:
        query = query.filter(Parcela.forma_pagamento_id == forma_pagamento_id)
    
    if data_inicio:
        query = query.filter(Parcela.data_pagamento >= dt_date.fromisoformat(data_inicio))
    
    if data_fim:
        query = query.filter(Parcela.data_pagamento <= dt_date.fromisoformat(data_fim))

    # Filtro por valor pago/valor (usa coalesce(valor_pago, valor))
    from sqlalchemy import func
    valor_expr = func.coalesce(Parcela.valor_pago, Parcela.valor)
    if valor_min is not None:
        query = query.filter(valor_expr >= valor_min)
    if valor_max is not None:
        query = query.filter(valor_expr <= valor_max)
    
    # Ordenar por data de pagamento (mais recentes primeiro)
    results = query.order_by(Parcela.data_pagamento.desc()).limit(limit).all()
    
    # Formatar resultados
    parcelas_pagas = []
    for r in results:
        parcelas_pagas.append({
            "id": r.id,
            "lancamento_id": r.lancamento_id,
            "numero_parcela": r.numero_parcela,
            "data_vencimento": r.data_vencimento.isoformat() if r.data_vencimento else None,
            "data_pagamento": r.data_pagamento.isoformat() if r.data_pagamento else None,
            "valor": float(r.valor),
            "valor_pago": float(r.valor_pago) if r.valor_pago else None,
            "tipo": r.tipo,
            "fornecedor": r.fornecedor,
            "forma_pagamento": {
                "id": r.forma_pagamento_id,
                "nome": r.forma_pagamento_nome,
                "tipo": r.forma_pagamento_tipo
            } if r.forma_pagamento_id else None,
            "observacao_pagamento": r.observacao_pagamento
        })
    
    return {
        "parcelas": parcelas_pagas,
        "total": len(parcelas_pagas)
    }

@app.get("/api/notificacoes")
async def obter_notificacoes(current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """
    Retorna notificações de parcelas vencidas e a vencer
    """
    from datetime import date as dt_date, timedelta
    
    hoje = dt_date.today()
    dias_antecedencia = 7  # Notificar 7 dias antes
    data_limite = hoje + timedelta(days=dias_antecedencia)
    
    # Buscar parcelas não pagas
    parcelas = db.query(
        Parcela.id,
        Parcela.numero_parcela,
        Parcela.data_vencimento,
        Parcela.valor,
        Lancamento.fornecedor,
        Lancamento.tipo,
        TipoLancamento.nome.label('tipo_nome')
    ).join(
        Lancamento, Parcela.lancamento_id == Lancamento.id
    ).outerjoin(
        TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
    ).filter(
        Parcela.paga == 0,
        Parcela.data_vencimento <= data_limite,
        Parcela.usuario_id == current_user.id
    ).order_by(
        Parcela.data_vencimento
    ).all()
    
    notificacoes = []
    
    for parcela in parcelas:
        dias_diferenca = (parcela.data_vencimento - hoje).days
        
        # Determinar tipo e prioridade da notificação
        if dias_diferenca < 0:
            tipo = "vencida"
            prioridade = "alta"
            mensagem = f"Parcela vencida há {abs(dias_diferenca)} dia(s)"
        elif dias_diferenca == 0:
            tipo = "vence_hoje"
            prioridade = "alta"
            mensagem = "Vence hoje"
        elif dias_diferenca <= 3:
            tipo = "vence_proximamente"
            prioridade = "media"
            mensagem = f"Vence em {dias_diferenca} dia(s)"
        else:
            tipo = "a_vencer"
            prioridade = "baixa"
            mensagem = f"Vence em {dias_diferenca} dia(s)"
        
        notificacoes.append({
            "id": parcela.id,
            "tipo": tipo,
            "prioridade": prioridade,
            "mensagem": mensagem,
            "titulo": f"{parcela.fornecedor} - Parcela {parcela.numero_parcela}",
            "valor": float(parcela.valor),
            "data_vencimento": parcela.data_vencimento.isoformat(),
            "dias_diferenca": dias_diferenca,
            "natureza": parcela.tipo,
            "tipo_lancamento": parcela.tipo_nome or "Sem tipo",
            "data_criacao": hoje.isoformat()
        })
    
    # Estatísticas
    total = len(notificacoes)
    vencidas = len([n for n in notificacoes if n["tipo"] == "vencida"])
    vence_hoje = len([n for n in notificacoes if n["tipo"] == "vence_hoje"])
    proximamente = len([n for n in notificacoes if n["tipo"] == "vence_proximamente"])
    
    return {
        "notificacoes": notificacoes,
        "total": total,
        "nao_lidas": total,  # Por enquanto, todas não lidas
        "stats": {
            "vencidas": vencidas,
            "vence_hoje": vence_hoje,
            "proximamente": proximamente,
            "a_vencer": total - vencidas - vence_hoje - proximamente
        }
    }

@app.get("/api/fluxo-caixa")
async def obter_fluxo_caixa(
    data_inicio: str,
    data_fim: str,
    saldo_inicial: Optional[float] = 0.0,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    from datetime import date as dt_date, timedelta
    from collections import defaultdict
    from decimal import Decimal
    
    try:
        inicio = dt_date.fromisoformat(data_inicio)
        fim = dt_date.fromisoformat(data_fim)
    except ValueError:
        raise HTTPException(status_code=400, detail="Datas inválidas")
    
    # Buscar todas as parcelas não pagas no período
    parcelas_query = db.query(
        Parcela.data_vencimento,
        Parcela.valor,
        Lancamento.tipo
    ).join(
        Lancamento, Parcela.lancamento_id == Lancamento.id
    ).filter(
        Parcela.paga == 0,
        Parcela.data_vencimento >= inicio,
        Parcela.data_vencimento <= fim,
        Parcela.usuario_id == current_user.id
    ).all()
    
    # Agrupar por data
    fluxo_por_data = defaultdict(lambda: {"receitas": Decimal(0), "despesas": Decimal(0)})
    
    for data_venc, valor, tipo in parcelas_query:
        if tipo == "receita":
            fluxo_por_data[data_venc]["receitas"] += Decimal(str(valor))
        else:
            fluxo_por_data[data_venc]["despesas"] += Decimal(str(valor))
    
    # Gerar série temporal dia a dia
    resultado = []
    saldo_acumulado = Decimal(str(saldo_inicial))
    data_atual = inicio
    
    while data_atual <= fim:
        fluxo = fluxo_por_data.get(data_atual, {"receitas": Decimal(0), "despesas": Decimal(0)})
        receitas = float(fluxo["receitas"])
        despesas = float(fluxo["despesas"])
        saldo_dia = receitas - despesas
        saldo_acumulado += Decimal(str(saldo_dia))
        
        resultado.append({
            "data": data_atual.isoformat(),
            "receitas": receitas,
            "despesas": despesas,
            "saldo_dia": saldo_dia,
            "saldo_acumulado": float(saldo_acumulado)
        })
        
        data_atual += timedelta(days=1)
    
    # Calcular totais
    total_receitas = sum(d["receitas"] for d in resultado)
    total_despesas = sum(d["despesas"] for d in resultado)
    
    return {
        "fluxo": resultado,
        "resumo": {
            "total_receitas": total_receitas,
            "total_despesas": total_despesas,
            "saldo_final": float(saldo_acumulado),
            "saldo_inicial": saldo_inicial
        }
    }

@app.put("/api/lancamentos/{lancamento_id}")
async def atualizar_lancamento(lancamento_id: int, lancamento: LancamentoIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    from datetime import date
    
    db_lancamento = db.query(Lancamento).filter(
        Lancamento.id == lancamento_id,
        Lancamento.usuario_id == current_user.id
    ).first()
    if not db_lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    
    # Validar o tipo_lancamento_id se fornecido
    if lancamento.tipo_lancamento_id:
        tipo = db.query(TipoLancamento).filter(TipoLancamento.id == lancamento.tipo_lancamento_id).first()
        if not tipo:
            raise HTTPException(status_code=404, detail="Tipo de lançamento não encontrado")
        if tipo.natureza != lancamento.tipo:
            raise HTTPException(status_code=400, detail="O tipo de lançamento selecionado não corresponde à natureza do lançamento")
    
    # Validar o subtipo_lancamento_id se fornecido
    if lancamento.subtipo_lancamento_id:
        subtipo = db.query(SubtipoLancamento).filter(SubtipoLancamento.id == lancamento.subtipo_lancamento_id).first()
        if not subtipo:
            raise HTTPException(status_code=404, detail="Subtipo de lançamento não encontrado")
        if not subtipo.ativo:
            raise HTTPException(status_code=400, detail=f"O subtipo '{subtipo.nome}' está inativo")
        if lancamento.tipo_lancamento_id and subtipo.tipo_lancamento_id != lancamento.tipo_lancamento_id:
            raise HTTPException(status_code=400, detail="O subtipo não pertence ao tipo selecionado")
    
    # Atualizar campos do lançamento
    db_lancamento.data_lancamento = date.fromisoformat(lancamento.data_lancamento)
    db_lancamento.tipo = lancamento.tipo
    db_lancamento.tipo_lancamento_id = lancamento.tipo_lancamento_id
    db_lancamento.subtipo_lancamento_id = lancamento.subtipo_lancamento_id
    db_lancamento.fornecedor = lancamento.fornecedor
    db_lancamento.valor_total = "{:.2f}".format(lancamento.valor_total)
    db_lancamento.data_primeiro_vencimento = date.fromisoformat(lancamento.data_primeiro_vencimento)
    db_lancamento.numero_parcelas = lancamento.numero_parcelas
    db_lancamento.valor_medio_parcelas = "{:.2f}".format(lancamento.valor_medio_parcelas)
    db_lancamento.observacao = lancamento.observacao

    try:
        # Excluir parcelas antigas
        db.query(Parcela).filter(Parcela.lancamento_id == lancamento_id).delete()
        db.commit()
        db.refresh(db_lancamento)

        # Recriar parcelas com distribuição de centavos
        from dateutil.relativedelta import relativedelta
        from decimal import Decimal
        data_venc = date.fromisoformat(lancamento.data_primeiro_vencimento)
        parcelas_criadas = []

        valor_total_decimal = Decimal(str(lancamento.valor_total))
        valor_parcela_base = valor_total_decimal / lancamento.numero_parcelas
        valor_parcela_arredondado = valor_parcela_base.quantize(Decimal('0.01'))
        soma_parcelas = valor_parcela_arredondado * lancamento.numero_parcelas
        diferenca = valor_total_decimal - soma_parcelas
        centavos_restantes = int(diferenca * 100)

        for i in range(lancamento.numero_parcelas):
            valor_parcela = valor_parcela_arredondado
            if (centavos_restantes != 0):
                parcelas_faltantes = lancamento.numero_parcelas - i
                if parcelas_faltantes <= abs(centavos_restantes):
                    ajuste = Decimal('0.01') if centavos_restantes > 0 else Decimal('-0.01')
                    valor_parcela += ajuste
                    centavos_restantes -= (1 if centavos_restantes > 0 else -1)

            parcela = Parcela(
                usuario_id=current_user.id,
                lancamento_id=db_lancamento.id,
                numero_parcela=i + 1,
                data_vencimento=data_venc + relativedelta(months=i),
                valor="{:.2f}".format(valor_parcela),
                paga=0,
                data_pagamento=None
            )
            db.add(parcela)
            parcelas_criadas.append(parcela)

        db.commit()

        # Atribuir para resposta
        db_lancamento._parcelas = parcelas_criadas
        return LancamentoOut.from_orm(db_lancamento, incluir_parcelas=True).model_dump()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar lançamento: {str(e)}")

@app.delete("/api/lancamentos/{lancamento_id}")
async def excluir_lancamento(lancamento_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    lancamento = db.query(Lancamento).filter(
        Lancamento.id == lancamento_id,
        Lancamento.usuario_id == current_user.id
    ).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    
    try:
        # Excluir parcelas associadas
        db.query(Parcela).filter(Parcela.lancamento_id == lancamento_id).delete()
        
        # Excluir lançamento
        db.delete(lancamento)
        db.commit()
        return {"status": "ok", "message": "Lançamento e parcelas excluídos com sucesso"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao excluir lançamento: {str(e)}")

class ParcelaPagamentoIn(BaseModel):
    paga: bool
    data_pagamento: Optional[str] = None
    valor_pago: Optional[float] = None
    forma_pagamento_id: Optional[int] = None
    observacao_pagamento: Optional[str] = Field(None, max_length=500)

class ParcelaEdicaoIn(BaseModel):
    data_vencimento: str
    valor: float

class LancamentoRecorrenteIn(BaseModel):
    tipo: str = Field(..., pattern="^(despesa|receita)$")
    tipo_lancamento_id: Optional[int] = None
    fornecedor: str = Field(..., min_length=1, max_length=255)
    valor_total: float = Field(..., gt=0)
    dia_vencimento: int = Field(..., ge=1, le=31)
    numero_parcelas: int = Field(..., gt=0)
    frequencia: str = Field(..., pattern="^(mensal|trimestral|anual)$")
    data_inicio: str = Field(..., description="YYYY-MM-DD")
    observacao: Optional[str] = None

class LancamentoRecorrenteOut(BaseModel):
    id: int
    tipo: str
    tipo_lancamento_id: Optional[int]
    fornecedor: str
    valor_total: float
    dia_vencimento: int
    numero_parcelas: int
    frequencia: str
    ativo: int
    data_inicio: str
    ultima_geracao: Optional[str]
    observacao: Optional[str]
    created_at: str
    
    model_config = {
        "from_attributes": True
    }
    
    @classmethod
    def from_orm(cls, obj):
        return cls(
            id=obj.id,
            tipo=obj.tipo,
            tipo_lancamento_id=obj.tipo_lancamento_id,
            fornecedor=obj.fornecedor,
            valor_total=float(obj.valor_total),
            dia_vencimento=obj.dia_vencimento,
            numero_parcelas=obj.numero_parcelas,
            frequencia=obj.frequencia,
            ativo=obj.ativo,
            data_inicio=obj.data_inicio.isoformat() if hasattr(obj.data_inicio, 'isoformat') else obj.data_inicio,
            ultima_geracao=obj.ultima_geracao.isoformat() if obj.ultima_geracao and hasattr(obj.ultima_geracao, 'isoformat') else obj.ultima_geracao,
            observacao=obj.observacao,
            created_at=obj.created_at.isoformat() if hasattr(obj.created_at, 'isoformat') else obj.created_at
        )

class MetaIn(BaseModel):
    ano: int = Field(..., ge=2020, le=2100)
    mes: int = Field(..., ge=1, le=12)
    tipo_lancamento_id: Optional[int] = None
    valor_planejado: float = Field(..., gt=0)
    descricao: Optional[str] = Field(None, max_length=500)

class MetaOut(BaseModel):
    id: int
    ano: int
    mes: int
    tipo_lancamento_id: Optional[int]
    valor_planejado: float
    descricao: Optional[str]
    created_at: str
    updated_at: Optional[str]
    
    # Campos extras calculados
    tipo_nome: Optional[str] = None
    tipo_natureza: Optional[str] = None
    valor_realizado: Optional[float] = None
    percentual_realizado: Optional[float] = None
    status: Optional[str] = None  # "dentro", "atencao", "excedido"
    
    model_config = {
        "from_attributes": True
    }

@app.patch("/api/parcelas/{parcela_id}/pagar")
async def marcar_parcela_paga(parcela_id: int, dados: ParcelaPagamentoIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    from datetime import date as dt_date
    
    parcela = db.query(Parcela).filter(
        Parcela.id == parcela_id,
        Parcela.usuario_id == current_user.id
    ).first()
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela não encontrada")
    
    # Validar forma de pagamento se fornecida
    if dados.forma_pagamento_id:
        forma = db.query(FormaPagamento).filter(
            FormaPagamento.id == dados.forma_pagamento_id,
            FormaPagamento.usuario_id == current_user.id
        ).first()
        if not forma:
            raise HTTPException(status_code=404, detail="Forma de pagamento não encontrada")
        if not forma.ativo:
            raise HTTPException(status_code=400, detail=f"A forma de pagamento '{forma.nome}' está inativa")
    
    try:
        parcela.paga = 1 if dados.paga else 0
        
        if dados.paga:
            # Se está marcando como paga
            if dados.data_pagamento:
                parcela.data_pagamento = dt_date.fromisoformat(dados.data_pagamento)
            else:
                parcela.data_pagamento = dt_date.today()
            
            # Se forneceu valor pago, usa ele, senão usa o valor original
            if dados.valor_pago is not None:
                parcela.valor_pago = "{:.2f}".format(dados.valor_pago)
            else:
                parcela.valor_pago = parcela.valor
            
            # Salvar forma de pagamento e observação
            parcela.forma_pagamento_id = dados.forma_pagamento_id
            parcela.observacao_pagamento = dados.observacao_pagamento
        else:
            # Se está desmarcando como paga
            parcela.data_pagamento = None
            parcela.valor_pago = None
            parcela.forma_pagamento_id = None
            parcela.observacao_pagamento = None
        
        db.commit()
        db.refresh(parcela)
        return ParcelaOut.from_orm(parcela)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar parcela: {str(e)}")

@app.put("/api/parcelas/{parcela_id}")
async def editar_parcela(parcela_id: int, dados: ParcelaEdicaoIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    from datetime import date as dt_date
    
    parcela = db.query(Parcela).filter(
        Parcela.id == parcela_id,
        Parcela.usuario_id == current_user.id
    ).first()
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela não encontrada")
    
    try:
        # Atualiza data de vencimento
        parcela.data_vencimento = dt_date.fromisoformat(dados.data_vencimento)
        
        # Atualiza valor
        parcela.valor = "{:.2f}".format(dados.valor)
        
        db.commit()
        db.refresh(parcela)
        return ParcelaOut.from_orm(parcela)
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Data inválida: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao editar parcela: {str(e)}")

# Endpoints de Lançamentos Recorrentes
@app.get("/api/recorrentes", response_model=List[LancamentoRecorrenteOut])
async def listar_recorrentes(current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    recorrentes = db.query(LancamentoRecorrente).filter(
        LancamentoRecorrente.usuario_id == current_user.id
    ).order_by(LancamentoRecorrente.fornecedor).all()
    return [LancamentoRecorrenteOut.from_orm(r) for r in recorrentes]

@app.post("/api/recorrentes", response_model=LancamentoRecorrenteOut)
async def criar_recorrente(recorrente: LancamentoRecorrenteIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    from datetime import date as dt_date
    
    # Validar tipo_lancamento_id se fornecido
    if recorrente.tipo_lancamento_id:
        tipo = db.query(TipoLancamento).filter(TipoLancamento.id == recorrente.tipo_lancamento_id).first()
        if not tipo:
            raise HTTPException(status_code=404, detail="Tipo de lançamento não encontrado")
        if tipo.natureza != recorrente.tipo:
            raise HTTPException(status_code=400, detail="O tipo de lançamento não corresponde à natureza")
    
    try:
        novo_recorrente = LancamentoRecorrente(
            usuario_id=current_user.id,
            tipo=recorrente.tipo,
            tipo_lancamento_id=recorrente.tipo_lancamento_id,
            fornecedor=recorrente.fornecedor,
            valor_total="{:.2f}".format(recorrente.valor_total),
            dia_vencimento=recorrente.dia_vencimento,
            numero_parcelas=recorrente.numero_parcelas,
            frequencia=recorrente.frequencia,
            ativo=1,
            data_inicio=dt_date.fromisoformat(recorrente.data_inicio),
            ultima_geracao=None,
            observacao=recorrente.observacao,
            created_at=dt_date.today()
        )
        
        db.add(novo_recorrente)
        db.commit()
        db.refresh(novo_recorrente)
        
        return LancamentoRecorrenteOut.from_orm(novo_recorrente)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao criar recorrente: {str(e)}")

@app.put("/api/recorrentes/{recorrente_id}", response_model=LancamentoRecorrenteOut)
async def atualizar_recorrente(recorrente_id: int, recorrente: LancamentoRecorrenteIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    from datetime import date as dt_date
    
    db_recorrente = db.query(LancamentoRecorrente).filter(
        LancamentoRecorrente.id == recorrente_id,
        LancamentoRecorrente.usuario_id == current_user.id
    ).first()
    if not db_recorrente:
        raise HTTPException(status_code=404, detail="Recorrente não encontrado")
    
    # Validar tipo_lancamento_id
    if recorrente.tipo_lancamento_id:
        tipo = db.query(TipoLancamento).filter(TipoLancamento.id == recorrente.tipo_lancamento_id).first()
        if not tipo:
            raise HTTPException(status_code=404, detail="Tipo de lançamento não encontrado")
        if tipo.natureza != recorrente.tipo:
            raise HTTPException(status_code=400, detail="O tipo de lançamento não corresponde à natureza")
    
    try:
        db_recorrente.tipo = recorrente.tipo
        db_recorrente.tipo_lancamento_id = recorrente.tipo_lancamento_id
        db_recorrente.fornecedor = recorrente.fornecedor
        db_recorrente.valor_total = "{:.2f}".format(recorrente.valor_total)
        db_recorrente.dia_vencimento = recorrente.dia_vencimento
        db_recorrente.numero_parcelas = recorrente.numero_parcelas
        db_recorrente.frequencia = recorrente.frequencia
        db_recorrente.data_inicio = dt_date.fromisoformat(recorrente.data_inicio)
        db_recorrente.observacao = recorrente.observacao
        
        db.commit()
        db.refresh(db_recorrente)
        
        return LancamentoRecorrenteOut.from_orm(db_recorrente)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar recorrente: {str(e)}")

@app.delete("/api/recorrentes/{recorrente_id}")
async def excluir_recorrente(recorrente_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    recorrente = db.query(LancamentoRecorrente).filter(
        LancamentoRecorrente.id == recorrente_id,
        LancamentoRecorrente.usuario_id == current_user.id
    ).first()
    if not recorrente:
        raise HTTPException(status_code=404, detail="Recorrente não encontrado")
    
    try:
        db.delete(recorrente)
        db.commit()
        return {"message": "Recorrente excluído com sucesso"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao excluir recorrente: {str(e)}")

@app.patch("/api/recorrentes/{recorrente_id}/toggle")
async def toggle_recorrente(recorrente_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    recorrente = db.query(LancamentoRecorrente).filter(
        LancamentoRecorrente.id == recorrente_id,
        LancamentoRecorrente.usuario_id == current_user.id
    ).first()
    if not recorrente:
        raise HTTPException(status_code=404, detail="Recorrente não encontrado")
    
    try:
        recorrente.ativo = 1 if recorrente.ativo == 0 else 0
        db.commit()
        db.refresh(recorrente)
        return LancamentoRecorrenteOut.from_orm(recorrente)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao alternar recorrente: {str(e)}")

@app.post("/api/recorrentes/{recorrente_id}/gerar")
async def gerar_lancamento_recorrente(recorrente_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    from datetime import date as dt_date
    from dateutil.relativedelta import relativedelta
    
    recorrente = db.query(LancamentoRecorrente).filter(
        LancamentoRecorrente.id == recorrente_id,
        LancamentoRecorrente.usuario_id == current_user.id
    ).first()
    if not recorrente:
        raise HTTPException(status_code=404, detail="Recorrente não encontrado")
    
    if recorrente.ativo == 0:
        raise HTTPException(status_code=400, detail="Recorrente está inativo")
    
    try:
        hoje = dt_date.today()
        
        # Calcular data do primeiro vencimento baseado na frequência
        if recorrente.frequencia == "mensal":
            # Próximo mês, no dia especificado
            data_venc = dt_date(hoje.year, hoje.month, min(recorrente.dia_vencimento, 28))
            if data_venc <= hoje:
                data_venc = data_venc + relativedelta(months=1)
        elif recorrente.frequencia == "trimestral":
            # Próximo trimestre
            data_venc = dt_date(hoje.year, hoje.month, min(recorrente.dia_vencimento, 28))
            if data_venc <= hoje:
                data_venc = data_venc + relativedelta(months=3)
        else:  # anual
            # Próximo ano
            data_venc = dt_date(hoje.year, hoje.month, min(recorrente.dia_vencimento, 28))
            if data_venc <= hoje:
                data_venc = data_venc + relativedelta(years=1)
        
        # Criar lançamento
        valor_medio = float(recorrente.valor_total) / recorrente.numero_parcelas
        
        novo_lancamento = Lancamento(
            usuario_id=current_user.id,
            data_lancamento=hoje,
            tipo=recorrente.tipo,
            tipo_lancamento_id=recorrente.tipo_lancamento_id,
            fornecedor=recorrente.fornecedor,
            valor_total=recorrente.valor_total,
            data_primeiro_vencimento=data_venc,
            numero_parcelas=recorrente.numero_parcelas,
            valor_medio_parcelas="{:.2f}".format(valor_medio),
            observacao=f"[AUTO-GERADO] {recorrente.observacao or ''}"
        )
        
        db.add(novo_lancamento)
        db.flush()
        
        # Gerar parcelas
        for i in range(recorrente.numero_parcelas):
            if recorrente.frequencia == "mensal":
                data_parcela = data_venc + relativedelta(months=i)
            elif recorrente.frequencia == "trimestral":
                data_parcela = data_venc + relativedelta(months=i*3)
            else:  # anual
                data_parcela = data_venc + relativedelta(years=i)
            
            parcela = Parcela(
                usuario_id=current_user.id,
                lancamento_id=novo_lancamento.id,
                numero_parcela=i + 1,
                data_vencimento=data_parcela,
                valor="{:.2f}".format(valor_medio),
                paga=0
            )
            db.add(parcela)
        
        # Atualizar ultima_geracao
        recorrente.ultima_geracao = hoje
        
        db.commit()
        db.refresh(novo_lancamento)
        
        return {
            "message": "Lançamento gerado com sucesso",
            "lancamento_id": novo_lancamento.id
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao gerar lançamento: {str(e)}")

@app.get("/api/dashboard")
async def obter_dashboard(
    tipo_data: Optional[str] = "vencimento",
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    natureza: Optional[str] = None,
    tipos: Optional[str] = None,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    from datetime import date, datetime, timedelta
    from sqlalchemy import func
    
    # Definir período padrão (mês atual)
    if not data_inicio:
        hoje = date.today()
        data_inicio = date(hoje.year, hoje.month, 1).isoformat()
    if not data_fim:
        hoje = date.today()
        proximo_mes = date(hoje.year + (1 if hoje.month == 12 else 0), 1 if hoje.month == 12 else hoje.month + 1, 1)
        ultimo_dia = proximo_mes - timedelta(days=1)
        data_fim = ultimo_dia.isoformat()
    
    # Processar filtro de tipos (lista de IDs separados por vírgula)
    tipos_ids = []
    if tipos:
        try:
            tipos_ids = [int(id.strip()) for id in tipos.split(',') if id.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="IDs de tipos inválidos")
    
    # Função para aplicar filtros adicionais
    def aplicar_filtros_adicionais(query):
        if natureza:
            query = query.filter(Lancamento.tipo == natureza)
        if tipos_ids:
            query = query.filter(Lancamento.tipo_lancamento_id.in_(tipos_ids))
        return query
    
    # Se filtrar por data de pagamento, consultar as PARCELAS pagas
    if tipo_data == "pagamento":
        # Totalizadores por parcelas PAGAS (usando valor_pago)
        query_receitas = db.query(func.sum(Parcela.valor_pago)).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "receita",
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_pagamento >= date.fromisoformat(data_inicio),
            Parcela.data_pagamento <= date.fromisoformat(data_fim)
        )
        query_receitas = aplicar_filtros_adicionais(query_receitas)
        total_receitas = query_receitas.scalar() or 0
        
        query_despesas = db.query(func.sum(Parcela.valor_pago)).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "despesa",
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_pagamento >= date.fromisoformat(data_inicio),
            Parcela.data_pagamento <= date.fromisoformat(data_fim)
        )
        query_despesas = aplicar_filtros_adicionais(query_despesas)
        total_despesas = query_despesas.scalar() or 0
        
        # Contar parcelas únicas (por lançamento)
        query_qtd_receitas = db.query(func.count(func.distinct(Parcela.lancamento_id))).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "receita",
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_pagamento >= date.fromisoformat(data_inicio),
            Parcela.data_pagamento <= date.fromisoformat(data_fim)
        )
        query_qtd_receitas = aplicar_filtros_adicionais(query_qtd_receitas)
        qtd_receitas = query_qtd_receitas.scalar() or 0
        
        query_qtd_despesas = db.query(func.count(func.distinct(Parcela.lancamento_id))).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "despesa",
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_pagamento >= date.fromisoformat(data_inicio),
            Parcela.data_pagamento <= date.fromisoformat(data_fim)
        )
        query_qtd_despesas = aplicar_filtros_adicionais(query_qtd_despesas)
        qtd_despesas = query_qtd_despesas.scalar() or 0
        
        # Despesas por tipo (somando parcelas pagas)
        from sqlalchemy import literal
        nome_tipo_coalesce = func.coalesce(TipoLancamento.nome, literal('Sem tipo'))
        query_despesas_tipo = db.query(
            nome_tipo_coalesce.label('nome'),
            func.sum(Parcela.valor_pago).label("total")
        ).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Lancamento.tipo == "despesa",
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_pagamento >= date.fromisoformat(data_inicio),
            Parcela.data_pagamento <= date.fromisoformat(data_fim)
        )
        query_despesas_tipo = aplicar_filtros_adicionais(query_despesas_tipo)
        despesas_por_tipo = query_despesas_tipo.group_by(nome_tipo_coalesce).all()
        
        # Receitas por tipo (somando parcelas pagas)
        query_receitas_tipo = db.query(
            nome_tipo_coalesce.label('nome'),
            func.sum(Parcela.valor_pago).label("total")
        ).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Lancamento.tipo == "receita",
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_pagamento >= date.fromisoformat(data_inicio),
            Parcela.data_pagamento <= date.fromisoformat(data_fim)
        )
        query_receitas_tipo = aplicar_filtros_adicionais(query_receitas_tipo)
        receitas_por_tipo = query_receitas_tipo.group_by(nome_tipo_coalesce).all()
    
    # Se filtrar por vencimento, consultar as PARCELAS
    elif tipo_data == "vencimento":
        # Totalizadores por parcelas
        query_rec_venc = db.query(func.sum(Parcela.valor)).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "receita",
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_vencimento >= date.fromisoformat(data_inicio),
            Parcela.data_vencimento <= date.fromisoformat(data_fim)
        )
        query_rec_venc = aplicar_filtros_adicionais(query_rec_venc)
        total_receitas = query_rec_venc.scalar() or 0
        
        query_desp_venc = db.query(func.sum(Parcela.valor)).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "despesa",
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_vencimento >= date.fromisoformat(data_inicio),
            Parcela.data_vencimento <= date.fromisoformat(data_fim)
        )
        query_desp_venc = aplicar_filtros_adicionais(query_desp_venc)
        total_despesas = query_desp_venc.scalar() or 0
        
        # Contar parcelas únicas (por lançamento)
        query_qtd_rec_venc = db.query(func.count(func.distinct(Parcela.lancamento_id))).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "receita",
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_vencimento >= date.fromisoformat(data_inicio),
            Parcela.data_vencimento <= date.fromisoformat(data_fim)
        )
        query_qtd_rec_venc = aplicar_filtros_adicionais(query_qtd_rec_venc)
        qtd_receitas = query_qtd_rec_venc.scalar() or 0
        
        query_qtd_desp_venc = db.query(func.count(func.distinct(Parcela.lancamento_id))).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Lancamento.tipo == "despesa",
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_vencimento >= date.fromisoformat(data_inicio),
            Parcela.data_vencimento <= date.fromisoformat(data_fim)
        )
        query_qtd_desp_venc = aplicar_filtros_adicionais(query_qtd_desp_venc)
        qtd_despesas = query_qtd_desp_venc.scalar() or 0
        
        # Despesas por tipo (somando parcelas)
        from sqlalchemy import literal
        nome_tipo_coalesce = func.coalesce(TipoLancamento.nome, literal('Sem tipo'))
        query_desp_tipo_venc = db.query(
            nome_tipo_coalesce.label('nome'),
            func.sum(Parcela.valor).label("total")
        ).select_from(Parcela).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Lancamento.tipo == "despesa",
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_vencimento >= date.fromisoformat(data_inicio),
            Parcela.data_vencimento <= date.fromisoformat(data_fim)
        )
        query_desp_tipo_venc = aplicar_filtros_adicionais(query_desp_tipo_venc)
        despesas_por_tipo = query_desp_tipo_venc.group_by(nome_tipo_coalesce).all()
        
        # Receitas por tipo (somando parcelas)
        query_rec_tipo_venc = db.query(
            nome_tipo_coalesce.label('nome'),
            func.sum(Parcela.valor).label("total")
        ).select_from(Parcela).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Lancamento.tipo == "receita",
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_vencimento >= date.fromisoformat(data_inicio),
            Parcela.data_vencimento <= date.fromisoformat(data_fim)
        )
        query_rec_tipo_venc = aplicar_filtros_adicionais(query_rec_tipo_venc)
        receitas_por_tipo = query_rec_tipo_venc.group_by(nome_tipo_coalesce).all()
    
    else:  # Filtrar por data de lançamento
        campo_data = Lancamento.data_lancamento
        
        query = db.query(Lancamento).filter(
            Lancamento.usuario_id == current_user.id,
            campo_data >= date.fromisoformat(data_inicio),
            campo_data <= date.fromisoformat(data_fim)
        )
        query = aplicar_filtros_adicionais(query)
        
        # Totalizadores
        query_rec_lanc = db.query(func.sum(Lancamento.valor_total)).filter(
            Lancamento.tipo == "receita",
            Lancamento.usuario_id == current_user.id,
            campo_data >= date.fromisoformat(data_inicio),
            campo_data <= date.fromisoformat(data_fim)
        )
        query_rec_lanc = aplicar_filtros_adicionais(query_rec_lanc)
        total_receitas = query_rec_lanc.scalar() or 0
        
        query_desp_lanc = db.query(func.sum(Lancamento.valor_total)).filter(
            Lancamento.tipo == "despesa",
            Lancamento.usuario_id == current_user.id,
            campo_data >= date.fromisoformat(data_inicio),
            campo_data <= date.fromisoformat(data_fim)
        )
        query_desp_lanc = aplicar_filtros_adicionais(query_desp_lanc)
        total_despesas = query_desp_lanc.scalar() or 0
        
        # Contar lançamentos
        qtd_receitas = query.filter(Lancamento.tipo == "receita").count()
        qtd_despesas = query.filter(Lancamento.tipo == "despesa").count()
        
        # Despesas por tipo
        from sqlalchemy import literal
        nome_tipo_coalesce = func.coalesce(TipoLancamento.nome, literal('Sem tipo'))
        query_desp_tipo_lanc = db.query(
            nome_tipo_coalesce.label('nome'),
            func.sum(Lancamento.valor_total).label("total")
        ).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Lancamento.tipo == "despesa",
            Lancamento.usuario_id == current_user.id,
            campo_data >= date.fromisoformat(data_inicio),
            campo_data <= date.fromisoformat(data_fim)
        )
        query_desp_tipo_lanc = aplicar_filtros_adicionais(query_desp_tipo_lanc)
        despesas_por_tipo = query_desp_tipo_lanc.group_by(nome_tipo_coalesce).all()
        
        # Receitas por tipo
        query_rec_tipo_lanc = db.query(
            nome_tipo_coalesce.label('nome'),
            func.sum(Lancamento.valor_total).label("total")
        ).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Lancamento.tipo == "receita",
            Lancamento.usuario_id == current_user.id,
            campo_data >= date.fromisoformat(data_inicio),
            campo_data <= date.fromisoformat(data_fim)
        )
        query_rec_tipo_lanc = aplicar_filtros_adicionais(query_rec_tipo_lanc)
        receitas_por_tipo = query_rec_tipo_lanc.group_by(nome_tipo_coalesce).all()
    
    return {
        "periodo": {
            "tipo_data": tipo_data,
            "data_inicio": data_inicio,
            "data_fim": data_fim
        },
        "totalizadores": {
            "receitas": float(total_receitas),
            "despesas": float(total_despesas),
            "saldo": float(total_receitas) - float(total_despesas),
            "qtd_receitas": qtd_receitas,
            "qtd_despesas": qtd_despesas
        },
        "despesas_por_tipo": [{"nome": nome, "total": float(total)} for nome, total in despesas_por_tipo],
        "receitas_por_tipo": [{"nome": nome, "total": float(total)} for nome, total in receitas_por_tipo]
    }

@app.get("/api/dashboard/tabela-anual")
async def obter_tabela_anual(
    ano: int,
    tipo_data: Optional[str] = "vencimento",
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """
    Retorna tabela anual com valores mensais agrupados por tipo de lançamento.
    tipo_data: 'vencimento' (padrão) ou 'pagamento'
    - vencimento: considera todas as parcelas pela data de vencimento
    - pagamento: considera apenas parcelas pagas pela data de pagamento
    """
    from datetime import date
    from sqlalchemy import func, extract
    
    # Definir campo de data e filtros baseado no tipo_data
    if tipo_data == "pagamento":
        # Usar data de pagamento, apenas parcelas pagas
        campo_data = Parcela.data_pagamento
        campo_valor = Parcela.valor_pago
        query = db.query(
            TipoLancamento.id,
            TipoLancamento.nome,
            TipoLancamento.natureza,
            extract('month', campo_data).label('mes'),
            func.sum(campo_valor).label('total')
        ).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).join(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            extract('year', campo_data) == ano
        )
    else:
        # Usar data de vencimento, todas as parcelas
        campo_data = Parcela.data_vencimento
        campo_valor = Parcela.valor
        query = db.query(
            TipoLancamento.id,
            TipoLancamento.nome,
            TipoLancamento.natureza,
            extract('month', campo_data).label('mes'),
            func.sum(campo_valor).label('total')
        ).join(
            Lancamento, Parcela.lancamento_id == Lancamento.id
        ).join(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).filter(
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            extract('year', campo_data) == ano
        )
    
    # Finalizar query
    query = query.group_by(
        TipoLancamento.id,
        TipoLancamento.nome,
        TipoLancamento.natureza,
        extract('month', campo_data)
    ).order_by(
        TipoLancamento.natureza.desc(),  # receitas primeiro
        TipoLancamento.nome
    )
    
    results = query.all()
    
    # Organizar dados por tipo
    tipos_dict = {}
    for tipo_id, nome, natureza, mes, total in results:
        if tipo_id not in tipos_dict:
            tipos_dict[tipo_id] = {
                "id": tipo_id,
                "nome": nome,
                "natureza": natureza,
                "meses": {}
            }
        tipos_dict[tipo_id]["meses"][int(mes)] = float(total)
    
    # Converter para lista
    tipos_list = list(tipos_dict.values())
    
    return {
        "ano": ano,
        "tipo_data": tipo_data,
        "tipos": tipos_list
    }

@app.get("/api/relatorios/tabela-anual-pdf")
async def exportar_tabela_anual_pdf(
    ano: int,
    tipo_data: Optional[str] = "vencimento",
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """
    Exporta a tabela anual em formato PDF.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    
    # Obter dados da tabela anual
    dados_tabela = await obter_tabela_anual(ano, tipo_data, current_user, db)
    
    if not dados_tabela["tipos"]:
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado para este ano")
    
    # Criar buffer para o PDF
    buffer = BytesIO()
    
    # Configurar documento PDF em paisagem (landscape)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    # Elementos do PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=12,
        alignment=1  # Center
    )
    
    tipo_data_label = "Pagamento" if tipo_data == "pagamento" else "Vencimento"
    titulo = f"Resumo Anual {ano} - Por Data de {tipo_data_label}"
    elements.append(Paragraph(titulo, title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Subtítulo com informação do tipo de data
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=20,
        alignment=1
    )
    
    info_text = "Valores baseados na data de pagamento (apenas parcelas pagas)" if tipo_data == "pagamento" else "Valores baseados na data de vencimento (todas as parcelas)"
    elements.append(Paragraph(info_text, subtitle_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Preparar dados da tabela
    meses = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    # Cabeçalho da tabela
    table_data = [['Tipo'] + meses + ['Total']]
    
    receitas = [t for t in dados_tabela["tipos"] if t["natureza"] == "receita"]
    despesas = [t for t in dados_tabela["tipos"] if t["natureza"] == "despesa"]
    
    def formatar_valor(valor):
        if valor == 0:
            return "-"
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    # Adicionar receitas
    if receitas:
        table_data.append(['RECEITAS'] + [''] * 13)
        for tipo in receitas:
            row = [tipo["nome"]]
            total_tipo = 0
            for mes in range(1, 13):
                valor = tipo["meses"].get(mes, 0)
                total_tipo += valor
                row.append(formatar_valor(valor))
            row.append(formatar_valor(total_tipo))
            table_data.append(row)
    
    # Adicionar despesas
    if despesas:
        table_data.append(['DESPESAS'] + [''] * 13)
        for tipo in despesas:
            row = [tipo["nome"]]
            total_tipo = 0
            for mes in range(1, 13):
                valor = tipo["meses"].get(mes, 0)
                total_tipo += valor
                row.append(formatar_valor(valor))
            row.append(formatar_valor(total_tipo))
            table_data.append(row)
    
    # Calcular totais mensais
    totais_row = ['TOTAL GERAL']
    total_geral = 0
    for mes in range(1, 13):
        total_mes = 0
        for tipo in dados_tabela["tipos"]:
            valor = tipo["meses"].get(mes, 0)
            if tipo["natureza"] == "receita":
                total_mes += valor
            else:
                total_mes -= valor
        total_geral += total_mes
        totais_row.append(formatar_valor(total_mes))
    totais_row.append(formatar_valor(total_geral))
    table_data.append(totais_row)
    
    # Criar tabela
    table = Table(table_data, repeatRows=1)
    
    # Estilo da tabela
    table_style = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22d3ee')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # Primeira coluna (nomes)
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (0, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        
        # Valores numéricos
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (1, 1), (-1, -1), 7),
        
        # Última coluna (Total) em destaque
        ('BACKGROUND', (-1, 1), (-1, -1), colors.HexColor('#f0f9ff')),
        ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ])
    
    # Destacar linhas de seção (RECEITAS, DESPESAS)
    row_idx = 1
    if receitas:
        table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#dcfce7'))
        table_style.add('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
        table_style.add('FONTSIZE', (0, row_idx), (-1, row_idx), 9)
        row_idx += len(receitas) + 1
    
    if despesas:
        table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#fee2e2'))
        table_style.add('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
        table_style.add('FONTSIZE', (0, row_idx), (-1, row_idx), 9)
    
    # Destacar linha de total geral
    table_style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0f2fe'))
    table_style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
    table_style.add('FONTSIZE', (0, -1), (-1, -1), 9)
    table_style.add('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#22d3ee'))
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Adicionar rodapé
    elements.append(Spacer(1, 0.5*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#94a3b8'),
        alignment=1
    )
    footer_text = f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} - Sistema Financeiro Pessoal"
    elements.append(Paragraph(footer_text, footer_style))
    
    # Gerar PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Retornar PDF como resposta
    filename = f"relatorio_anual_{ano}_{tipo_data}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@app.get("/api/relatorios/lancamentos-excel")
async def exportar_lancamentos_excel(
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    tipo: Optional[str] = None,
    tipo_lancamento_id: Optional[int] = None,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """
    Exporta lançamentos financeiros em formato Excel (.xlsx).
    Permite filtrar por período, tipo (receita/despesa) e tipo de lançamento.
    """
    import pandas as pd
    from io import BytesIO
    from datetime import datetime
    
    # Query base
    query = db.query(
        Lancamento.id,
        Lancamento.data_lancamento,
        Lancamento.tipo,
        Lancamento.fornecedor,
        Lancamento.valor_total,
        Lancamento.numero_parcelas,
        Lancamento.valor_medio_parcelas,
        Lancamento.data_primeiro_vencimento,
        Lancamento.observacao,
        TipoLancamento.nome.label('tipo_lancamento_nome'),
        SubtipoLancamento.nome.label('subtipo_lancamento_nome')
    ).outerjoin(
        TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
    ).outerjoin(
        SubtipoLancamento, Lancamento.subtipo_lancamento_id == SubtipoLancamento.id
    ).filter(
        Lancamento.usuario_id == current_user.id
    )
    
    # Aplicar filtros
    if data_inicio:
        query = query.filter(Lancamento.data_lancamento >= data_inicio)
    if data_fim:
        query = query.filter(Lancamento.data_lancamento <= data_fim)
    if tipo:
        query = query.filter(Lancamento.tipo == tipo)
    if tipo_lancamento_id:
        query = query.filter(Lancamento.tipo_lancamento_id == tipo_lancamento_id)
    
    query = query.order_by(Lancamento.data_lancamento.desc())
    
    lancamentos = query.all()
    
    if not lancamentos:
        raise HTTPException(status_code=404, detail="Nenhum lançamento encontrado com os filtros aplicados")
    
    # Converter para DataFrame
    data = []
    for lanc in lancamentos:
        data.append({
            'ID': lanc.id,
            'Data Lançamento': lanc.data_lancamento.strftime('%d/%m/%Y'),
            'Tipo': '📈 Receita' if lanc.tipo == 'receita' else '📉 Despesa',
            'Categoria': lanc.tipo_lancamento_nome or 'Sem categoria',
            'Subcategoria': lanc.subtipo_lancamento_nome or '-',
            'Fornecedor/Cliente': lanc.fornecedor,
            'Valor Total': float(lanc.valor_total),
            'Nº Parcelas': lanc.numero_parcelas,
            'Valor Médio Parcela': float(lanc.valor_medio_parcelas),
            'Primeiro Vencimento': lanc.data_primeiro_vencimento.strftime('%d/%m/%Y'),
            'Observação': lanc.observacao or ''
        })
    
    df = pd.DataFrame(data)
    
    # Criar arquivo Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Lançamentos', index=False)
        
        # Formatar planilha
        workbook = writer.book
        worksheet = writer.sheets['Lançamentos']
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 8,   # ID
            'B': 18,  # Data Lançamento
            'C': 15,  # Tipo
            'D': 25,  # Categoria
            'E': 20,  # Subcategoria
            'F': 30,  # Fornecedor
            'G': 15,  # Valor Total
            'H': 12,  # Nº Parcelas
            'I': 18,  # Valor Médio
            'J': 18,  # Primeiro Venc
            'K': 40   # Observação
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatar valores monetários
        from openpyxl.styles import numbers
        for row in range(2, len(df) + 2):
            worksheet[f'G{row}'].number_format = 'R$ #,##0.00'
            worksheet[f'I{row}'].number_format = 'R$ #,##0.00'
        
        # Estilo do cabeçalho
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color='22D3EE', end_color='22D3EE', fill_type='solid')
        header_font = Font(bold=True, color='0F172A')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"lancamentos_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@app.get("/api/relatorios/parcelas-excel")
async def exportar_parcelas_excel(
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    status: Optional[str] = None,  # "todas", "pagas", "pendentes"
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """
    Exporta parcelas em formato Excel (.xlsx).
    Permite filtrar por período e status de pagamento.
    """
    import pandas as pd
    from io import BytesIO
    from datetime import datetime
    
    # Query base
    query = db.query(
        Parcela.id,
        Parcela.numero_parcela,
        Parcela.data_vencimento,
        Parcela.valor,
        Parcela.paga,
        Parcela.data_pagamento,
        Parcela.valor_pago,
        Parcela.forma_pagamento_id,
        Lancamento.fornecedor,
        Lancamento.tipo,
        TipoLancamento.nome.label('tipo_lancamento_nome'),
        FormaPagamento.nome.label('forma_pagamento_nome')
    ).join(
        Lancamento, Parcela.lancamento_id == Lancamento.id
    ).outerjoin(
        TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
    ).outerjoin(
        FormaPagamento, Parcela.forma_pagamento_id == FormaPagamento.id
    ).filter(
        Parcela.usuario_id == current_user.id
    )
    
    # Aplicar filtros
    if data_inicio:
        query = query.filter(Parcela.data_vencimento >= data_inicio)
    if data_fim:
        query = query.filter(Parcela.data_vencimento <= data_fim)
    if status == "pagas":
        query = query.filter(Parcela.paga == 1)
    elif status == "pendentes":
        query = query.filter(Parcela.paga == 0)
    
    query = query.order_by(Parcela.data_vencimento.asc())
    
    parcelas = query.all()
    
    if not parcelas:
        raise HTTPException(status_code=404, detail="Nenhuma parcela encontrada com os filtros aplicados")
    
    # Converter para DataFrame
    data = []
    for parc in parcelas:
        data.append({
            'ID': parc.id,
            'Fornecedor/Cliente': parc.fornecedor,
            'Tipo': '📈 Receita' if parc.tipo == 'receita' else '📉 Despesa',
            'Categoria': parc.tipo_lancamento_nome or 'Sem categoria',
            'Nº Parcela': parc.numero_parcela,
            'Data Vencimento': parc.data_vencimento.strftime('%d/%m/%Y'),
            'Valor': float(parc.valor),
            'Status': '✅ Paga' if parc.paga else '⏳ Pendente',
            'Data Pagamento': parc.data_pagamento.strftime('%d/%m/%Y') if parc.data_pagamento else '',
            'Valor Pago': float(parc.valor_pago) if parc.valor_pago else 0.0,
            'Forma de Pagamento': parc.forma_pagamento_nome or ''
        })
    
    df = pd.DataFrame(data)
    
    # Criar arquivo Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Parcelas', index=False)
        
        # Formatar planilha
        workbook = writer.book
        worksheet = writer.sheets['Parcelas']
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 8,   # ID
            'B': 30,  # Fornecedor
            'C': 15,  # Tipo
            'D': 25,  # Categoria
            'E': 12,  # Nº Parcela
            'F': 18,  # Data Vencimento
            'G': 15,  # Valor
            'H': 15,  # Status
            'I': 18,  # Data Pagamento
            'J': 15,  # Valor Pago
            'K': 25   # Forma de Pagamento
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatar valores monetários
        from openpyxl.styles import numbers
        for row in range(2, len(df) + 2):
            worksheet[f'G{row}'].number_format = 'R$ #,##0.00'
            worksheet[f'J{row}'].number_format = 'R$ #,##0.00'
        
        # Estilo do cabeçalho
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color='22D3EE', end_color='22D3EE', fill_type='solid')
        header_font = Font(bold=True, color='0F172A')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"parcelas_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@app.get("/api/dashboard/evolucao")
async def obter_evolucao_mensal(
    meses: int = 6,
    tipo_data: Optional[str] = "pagamento",
    natureza: Optional[str] = None,
    tipos: Optional[str] = None,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """
    Retorna evolução mensal de receitas e despesas para os últimos N meses.
    - tipo_data: 'pagamento' (padrão), 'vencimento' ou 'lancamento'
    - natureza: filtra por 'receita' ou 'despesa' (opcional)
    - tipos: lista de ids separados por vírgula (opcional)
    """
    from datetime import date
    from sqlalchemy import func, extract

    if meses < 1:
        meses = 1
    if meses > 24:
        meses = 24

    # Processar filtro de tipos (lista de IDs separados por vírgula)
    tipos_ids: list[int] = []
    if tipos:
        try:
            tipos_ids = [int(id.strip()) for id in tipos.split(',') if id.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="IDs de tipos inválidos")

    hoje = date.today()
    # Gerar lista de (ano, mes) para os últimos N meses (mais antigo -> mais recente)
    labels = []
    anos_meses = []
    ano = hoje.year
    mes = hoje.month
    for _ in range(meses):
        anos_meses.append((ano, mes))
        labels.append(f"{ano}-{mes:02d}")
        # Decrementar um mês
        mes -= 1
        if mes == 0:
            mes = 12
            ano -= 1
    anos_meses.reverse()
    labels.reverse()

    # Função auxiliar para construir query base
    def build_query(natureza_alvo: str):
        if tipo_data == "pagamento":
            campo_data = Parcela.data_pagamento
            campo_valor = Parcela.valor_pago
            q = db.query(
                extract('year', campo_data).label('ano'),
                extract('month', campo_data).label('mes'),
                func.sum(campo_valor).label('total')
            ).join(
                Lancamento, Parcela.lancamento_id == Lancamento.id
            ).filter(
                Parcela.paga == 1,
                Lancamento.tipo == natureza_alvo,
                Parcela.usuario_id == current_user.id,
                extract('year', campo_data) * 100 + extract('month', campo_data)
                >= (anos_meses[0][0] * 100 + anos_meses[0][1]),
                extract('year', campo_data) * 100 + extract('month', campo_data)
                <= (anos_meses[-1][0] * 100 + anos_meses[-1][1])
            )
            if tipos_ids:
                q = q.filter(Lancamento.tipo_lancamento_id.in_(tipos_ids))
        elif tipo_data == "vencimento":
            campo_data = Parcela.data_vencimento
            campo_valor = Parcela.valor
            q = db.query(
                extract('year', campo_data).label('ano'),
                extract('month', campo_data).label('mes'),
                func.sum(campo_valor).label('total')
            ).join(
                Lancamento, Parcela.lancamento_id == Lancamento.id
            ).filter(
                Lancamento.tipo == natureza_alvo,
                Parcela.usuario_id == current_user.id,
                extract('year', campo_data) * 100 + extract('month', campo_data)
                >= (anos_meses[0][0] * 100 + anos_meses[0][1]),
                extract('year', campo_data) * 100 + extract('month', campo_data)
                <= (anos_meses[-1][0] * 100 + anos_meses[-1][1])
            )
            if tipos_ids:
                q = q.filter(Lancamento.tipo_lancamento_id.in_(tipos_ids))
        else:  # lancamento
            campo_data = Lancamento.data_lancamento
            campo_valor = Lancamento.valor_total
            q = db.query(
                extract('year', campo_data).label('ano'),
                extract('month', campo_data).label('mes'),
                func.sum(campo_valor).label('total')
            ).filter(
                Lancamento.tipo == natureza_alvo,
                Lancamento.usuario_id == current_user.id,
                extract('year', campo_data) * 100 + extract('month', campo_data)
                >= (anos_meses[0][0] * 100 + anos_meses[0][1]),
                extract('year', campo_data) * 100 + extract('month', campo_data)
                <= (anos_meses[-1][0] * 100 + anos_meses[-1][1])
            )
            if tipos_ids:
                q = q.filter(Lancamento.tipo_lancamento_id.in_(tipos_ids))
        return q.group_by('ano', 'mes').order_by('ano', 'mes')

    # Executar queries
    receitas_rows = []
    despesas_rows = []
    if not natureza or natureza == "receita":
        receitas_rows = build_query("receita").all()
    if not natureza or natureza == "despesa":
        despesas_rows = build_query("despesa").all()

    # Mapear resultados (ano,mes) -> total
    rec_map = {(int(a), int(m)): float(t or 0) for a, m, t in receitas_rows}
    desp_map = {(int(a), int(m)): float(t or 0) for a, m, t in despesas_rows}

    receitas_series = []
    despesas_series = []
    for (a, m) in anos_meses:
        receitas_series.append(rec_map.get((a, m), 0.0))
        despesas_series.append(desp_map.get((a, m), 0.0))

    return {
        "labels": [f"{a}-{m:02d}" for (a, m) in anos_meses],
        "receitas": receitas_series,
        "despesas": despesas_series,
        "tipo_data": tipo_data
    }

@app.get("/api/dashboard/top-formas")
async def obter_top_formas_pagamento(
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    limit: int = 3,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """
    Retorna as formas de pagamento mais usadas no período com variação M/M.
    
    - data_inicio/data_fim: período de análise (YYYY-MM-DD)
    - limit: número de formas a retornar (padrão 3)
    
    Retorna:
    {
        "top_formas": [
            {
                "forma_id": 1,
                "forma_nome": "Nubank",
                "total_pago": 5000.50,
                "quantidade_pagamentos": 25,
                "variacao_mes_anterior": 15.5  # Percentual de variação
            },
            ...
        ],
        "periodo": {
            "data_inicio": "2025-10-01",
            "data_fim": "2025-10-31"
        }
    }
    """
    from sqlalchemy import func, extract
    from datetime import date, timedelta
    from dateutil.relativedelta import relativedelta
    
    # Definir período padrão (mês atual)
    if not data_fim:
        data_fim = date.today().isoformat()
    if not data_inicio:
        fim_date = date.fromisoformat(data_fim)
        inicio_date = fim_date.replace(day=1)
        data_inicio = inicio_date.isoformat()
    else:
        inicio_date = date.fromisoformat(data_inicio)
        fim_date = date.fromisoformat(data_fim)
    
    # Calcular período anterior para comparação
    mes_anterior_inicio = inicio_date - relativedelta(months=1)
    mes_anterior_fim = fim_date - relativedelta(months=1)
    
    # Query principal: formas mais usadas no período atual
    top_formas_query = db.query(
        FormaPagamento.id.label('forma_id'),
        FormaPagamento.nome.label('forma_nome'),
        func.sum(Parcela.valor_pago).label('total_pago'),
        func.count(Parcela.id).label('quantidade_pagamentos')
    ).join(
        Parcela, FormaPagamento.id == Parcela.forma_pagamento_id
    ).filter(
        Parcela.paga == 1,
        Parcela.data_pagamento >= data_inicio,
        Parcela.data_pagamento <= data_fim,
        Parcela.usuario_id == current_user.id,
        FormaPagamento.usuario_id == current_user.id
    ).group_by(
        FormaPagamento.id, FormaPagamento.nome
    ).order_by(
        func.sum(Parcela.valor_pago).desc()
    ).limit(limit)
    
    top_formas = top_formas_query.all()
    
    # Calcular variação para cada forma
    resultado = []
    for forma in top_formas:
        # Valor do mês anterior
        valor_mes_anterior = db.query(
            func.sum(Parcela.valor_pago)
        ).join(
            FormaPagamento, FormaPagamento.id == Parcela.forma_pagamento_id
        ).filter(
            Parcela.paga == 1,
            Parcela.forma_pagamento_id == forma.forma_id,
            Parcela.data_pagamento >= mes_anterior_inicio.isoformat(),
            Parcela.data_pagamento <= mes_anterior_fim.isoformat(),
            Parcela.usuario_id == current_user.id,
            FormaPagamento.usuario_id == current_user.id
        ).scalar() or 0
        
        # Calcular variação percentual
        variacao = 0
        if valor_mes_anterior > 0:
            variacao = ((float(forma.total_pago) - float(valor_mes_anterior)) / float(valor_mes_anterior)) * 100
        elif forma.total_pago > 0:
            variacao = 100  # Novo (não existia no mês anterior)
        
        resultado.append({
            "forma_id": forma.forma_id,
            "forma_nome": forma.forma_nome,
            "total_pago": float(forma.total_pago),
            "quantidade_pagamentos": forma.quantidade_pagamentos,
            "variacao_mes_anterior": round(variacao, 1)
        })
    
    return {
        "top_formas": resultado,
        "periodo": {
            "data_inicio": data_inicio,
            "data_fim": data_fim
        }
    }

@app.get("/api/dashboard/por-tipo-subtipo")
async def obter_analise_hierarquica(
    tipo_data: Optional[str] = "vencimento",
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    natureza: Optional[str] = None,
    current_user: User = Depends(ensure_subscription),
    db: Session = Depends(get_db)
):
    """
    Retorna análise hierárquica de gastos/receitas por tipo e subtipo.
    Ideal para visualização em árvore expansível.
    
    - tipo_data: 'vencimento' (padrão), 'pagamento' ou 'lancamento'
    - data_inicio/data_fim: período de análise (YYYY-MM-DD)
    - natureza: 'receita', 'despesa' ou None (ambas)
    
    Retorna estrutura hierárquica:
    {
        "tipos": [
            {
                "tipo_id": 1,
                "tipo_nome": "Alimentação",
                "tipo_natureza": "despesa",
                "total": 1250.50,
                "quantidade_lancamentos": 15,
                "subtipos": [
                    {
                        "subtipo_id": 5,
                        "subtipo_nome": "Restaurante",
                        "total": 650.00,
                        "quantidade_lancamentos": 8,
                        "percentual_do_tipo": 52.0
                    },
                    ...
                ],
                "sem_subtipo": {
                    "total": 150.50,
                    "quantidade_lancamentos": 2,
                    "percentual_do_tipo": 12.0
                }
            },
            ...
        ],
        "totais": {
            "receitas": 5000.00,
            "despesas": 3500.00,
            "saldo": 1500.00
        }
    }
    """
    from datetime import date
    from sqlalchemy import func, case
    
    # Definir período padrão (mês atual)
    if not data_inicio:
        hoje = date.today()
        data_inicio = date(hoje.year, hoje.month, 1).isoformat()
    if not data_fim:
        hoje = date.today()
        proximo_mes = date(hoje.year + (1 if hoje.month == 12 else 0), 1 if hoje.month == 12 else hoje.month + 1, 1)
        ultimo_dia = proximo_mes - timedelta(days=1)
        data_fim = ultimo_dia.isoformat()
    
    data_inicio_obj = date.fromisoformat(data_inicio)
    data_fim_obj = date.fromisoformat(data_fim)
    
    # Determinar a query base conforme tipo_data
    if tipo_data == "pagamento":
        # Agrupar por tipo e subtipo via parcelas pagas
        query = db.query(
            TipoLancamento.id.label('tipo_id'),
            TipoLancamento.nome.label('tipo_nome'),
            TipoLancamento.natureza.label('tipo_natureza'),
            Lancamento.subtipo_lancamento_id.label('subtipo_id'),
            SubtipoLancamento.nome.label('subtipo_nome'),
            func.sum(Parcela.valor_pago).label('total'),
            func.count(func.distinct(Lancamento.id)).label('quantidade')
        ).select_from(Lancamento).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).outerjoin(
            SubtipoLancamento, Lancamento.subtipo_lancamento_id == SubtipoLancamento.id
        ).join(
            Parcela, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Parcela.paga == 1,
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_pagamento >= data_inicio_obj,
            Parcela.data_pagamento <= data_fim_obj
        )
    
    elif tipo_data == "vencimento":
        # Agrupar por tipo e subtipo via parcelas (vencimento)
        query = db.query(
            TipoLancamento.id.label('tipo_id'),
            TipoLancamento.nome.label('tipo_nome'),
            TipoLancamento.natureza.label('tipo_natureza'),
            Lancamento.subtipo_lancamento_id.label('subtipo_id'),
            SubtipoLancamento.nome.label('subtipo_nome'),
            func.sum(Parcela.valor).label('total'),
            func.count(func.distinct(Lancamento.id)).label('quantidade')
        ).select_from(Lancamento).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).outerjoin(
            SubtipoLancamento, Lancamento.subtipo_lancamento_id == SubtipoLancamento.id
        ).join(
            Parcela, Parcela.lancamento_id == Lancamento.id
        ).filter(
            Parcela.usuario_id == current_user.id,
            Lancamento.usuario_id == current_user.id,
            Parcela.data_vencimento >= data_inicio_obj,
            Parcela.data_vencimento <= data_fim_obj
        )
    
    else:  # lancamento
        # Agrupar por tipo e subtipo via lançamentos
        query = db.query(
            TipoLancamento.id.label('tipo_id'),
            TipoLancamento.nome.label('tipo_nome'),
            TipoLancamento.natureza.label('tipo_natureza'),
            Lancamento.subtipo_lancamento_id.label('subtipo_id'),
            SubtipoLancamento.nome.label('subtipo_nome'),
            func.sum(Lancamento.valor_total).label('total'),
            func.count(Lancamento.id).label('quantidade')
        ).select_from(Lancamento).outerjoin(
            TipoLancamento, Lancamento.tipo_lancamento_id == TipoLancamento.id
        ).outerjoin(
            SubtipoLancamento, Lancamento.subtipo_lancamento_id == SubtipoLancamento.id
        ).filter(
            Lancamento.usuario_id == current_user.id,
            Lancamento.data_lancamento >= data_inicio_obj,
            Lancamento.data_lancamento <= data_fim_obj
        )
    
    # Filtrar por natureza se especificado
    if natureza:
        if natureza not in ['receita', 'despesa']:
            raise HTTPException(status_code=400, detail="Natureza deve ser 'receita' ou 'despesa'")
        query = query.filter(Lancamento.tipo == natureza)
    
    # Agrupar e ordenar
    query = query.group_by(
        TipoLancamento.id,
        TipoLancamento.nome,
        TipoLancamento.natureza,
        Lancamento.subtipo_lancamento_id,
        SubtipoLancamento.nome
    ).order_by(
        TipoLancamento.natureza.desc(),  # despesas primeiro
        TipoLancamento.nome,
        SubtipoLancamento.nome
    )
    
    resultados = query.all()
    
    # Organizar em estrutura hierárquica
    tipos_map = {}
    total_receitas = 0.0
    total_despesas = 0.0
    
    for row in resultados:
        tipo_id = row.tipo_id
        tipo_nome = row.tipo_nome or "Sem tipo"
        tipo_natureza = row.tipo_natureza
        subtipo_id = row.subtipo_id
        subtipo_nome = row.subtipo_nome
        total = float(row.total or 0)
        quantidade = int(row.quantidade or 0)
        
        # Acumular totais gerais
        if tipo_natureza == 'receita':
            total_receitas += total
        else:
            total_despesas += total
        
        # Criar tipo se não existir
        if tipo_id not in tipos_map:
            tipos_map[tipo_id] = {
                "tipo_id": tipo_id,
                "tipo_nome": tipo_nome,
                "tipo_natureza": tipo_natureza,
                "total": 0.0,
                "quantidade_lancamentos": 0,
                "subtipos": [],
                "sem_subtipo": {
                    "total": 0.0,
                    "quantidade_lancamentos": 0,
                    "percentual_do_tipo": 0.0
                }
            }
        
        # Adicionar ao total do tipo
        tipos_map[tipo_id]["total"] += total
        tipos_map[tipo_id]["quantidade_lancamentos"] += quantidade
        
        # Separar lançamentos com e sem subtipo
        if subtipo_id:
            tipos_map[tipo_id]["subtipos"].append({
                "subtipo_id": subtipo_id,
                "subtipo_nome": subtipo_nome,
                "total": total,
                "quantidade_lancamentos": quantidade,
                "percentual_do_tipo": 0.0  # será calculado depois
            })
        else:
            tipos_map[tipo_id]["sem_subtipo"]["total"] += total
            tipos_map[tipo_id]["sem_subtipo"]["quantidade_lancamentos"] += quantidade
    
    # Calcular percentuais
    for tipo in tipos_map.values():
        total_tipo = tipo["total"]
        if total_tipo > 0:
            # Percentuais dos subtipos
            for subtipo in tipo["subtipos"]:
                subtipo["percentual_do_tipo"] = round((subtipo["total"] / total_tipo) * 100, 1)
            
            # Percentual do sem_subtipo
            if tipo["sem_subtipo"]["total"] > 0:
                tipo["sem_subtipo"]["percentual_do_tipo"] = round(
                    (tipo["sem_subtipo"]["total"] / total_tipo) * 100, 1
                )
    
    # Converter para lista e ordenar
    tipos_lista = list(tipos_map.values())
    tipos_lista.sort(key=lambda t: (
        0 if t["tipo_natureza"] == "despesa" else 1,  # despesas primeiro
        -t["total"]  # maior valor primeiro
    ))
    
    return {
        "tipos": tipos_lista,
        "totais": {
            "receitas": round(total_receitas, 2),
            "despesas": round(total_despesas, 2),
            "saldo": round(total_receitas - total_despesas, 2)
        },
        "filtros": {
            "tipo_data": tipo_data,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "natureza": natureza
        }
    }

# ========== ENDPOINTS DE BACKUP E VALIDAÇÃO ==========

@app.post("/api/backup/criar")
async def endpoint_criar_backup():
    """Cria um novo backup do banco de dados"""
    resultado = criar_backup()
    
    if not resultado.get("success"):
        raise HTTPException(status_code=500, detail=resultado.get("error"))
    
    # Limpar backups antigos (mais de 30 dias)
    removidos = limpar_backups_antigos(30)
    resultado["backups_removidos"] = len(removidos)
    
    return resultado

@app.get("/api/backup/listar")
async def endpoint_listar_backups():
    """Lista todos os backups disponíveis"""
    backups = listar_backups()
    return {
        "total": len(backups),
        "backups": backups
    }

@app.post("/api/backup/restaurar/{filename}")
async def endpoint_restaurar_backup(filename: str):
    """Restaura banco de dados de um backup específico"""
    resultado = restaurar_backup(filename)
    
    if not resultado.get("success"):
        raise HTTPException(status_code=500, detail=resultado.get("error"))
    
    return resultado

@app.delete("/api/backup/remover/{filename}")
async def endpoint_remover_backup(filename: str):
    """Remove um backup específico"""
    try:
        backup_path = BACKUP_DIR / filename
        
        if not backup_path.exists():
            raise HTTPException(status_code=404, detail="Backup não encontrado")
        
        backup_path.unlink()
        
        return {
            "success": True,
            "message": f"Backup {filename} removido com sucesso"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/backup/download/{filename}")
async def endpoint_download_backup(filename: str):
    """Faz download de um backup específico"""
    backup_path = BACKUP_DIR / filename
    
    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="Backup não encontrado")
    
    return FileResponse(
        path=str(backup_path),
        filename=filename,
        media_type="application/octet-stream"
    )

@app.get("/api/exportar/json")
async def endpoint_exportar_json(db: Session = Depends(get_db)):
    """Exporta todos os dados em formato JSON"""
    dados = exportar_dados_json(db)
    
    # Salvar arquivo JSON no diretório de backups
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = f"export_{timestamp}.json"
    json_path = BACKUP_DIR / json_filename
    
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(dados, f, indent=2, ensure_ascii=False)
    
    return {
        "success": True,
        "data": dados,
        "saved_to": json_filename,
        "path": str(json_path)
    }

@app.get("/api/diagnostico")
async def endpoint_diagnostico(db: Session = Depends(get_db)):
    """Executa diagnóstico de integridade dos dados"""
    resultado = validar_integridade(db)
    return resultado

# ========== ENDPOINTS DE METAS E ORÇAMENTO ==========

@app.post("/api/metas", response_model=MetaOut)
async def criar_meta(meta: MetaIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Cria uma nova meta mensal"""
    from datetime import date as dt_date
    
    # Verificar se já existe meta para este mês/ano/tipo
    meta_existente = db.query(Meta).filter(
        Meta.ano == meta.ano,
        Meta.mes == meta.mes,
        Meta.tipo_lancamento_id == meta.tipo_lancamento_id
    ).first()
    
    if meta_existente:
        raise HTTPException(
            status_code=400, 
            detail=f"Já existe uma meta para {meta.mes}/{meta.ano} neste tipo de lançamento"
        )
    
    nova_meta = Meta(
        ano=meta.ano,
        mes=meta.mes,
        tipo_lancamento_id=meta.tipo_lancamento_id,
        valor_planejado=meta.valor_planejado,
        descricao=meta.descricao,
        created_at=dt_date.today()
    )
    
    db.add(nova_meta)
    db.commit()
    db.refresh(nova_meta)
    
    # Carregar informações do tipo
    meta_out = MetaOut.model_validate(nova_meta)
    if nova_meta.tipo_lancamento_id:
        tipo = db.query(TipoLancamento).filter(TipoLancamento.id == nova_meta.tipo_lancamento_id).first()
        if tipo:
            meta_out.tipo_nome = tipo.nome
            meta_out.tipo_natureza = tipo.natureza
    
    return meta_out

@app.get("/api/metas")
async def listar_metas(
    ano: Optional[int] = None,
    mes: Optional[int] = None,
    tipo_lancamento_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Lista metas com filtros opcionais"""
    query = db.query(Meta)
    
    if ano:
        query = query.filter(Meta.ano == ano)
    if mes:
        query = query.filter(Meta.mes == mes)
    if tipo_lancamento_id:
        query = query.filter(Meta.tipo_lancamento_id == tipo_lancamento_id)
    
    metas = query.order_by(Meta.ano.desc(), Meta.mes.desc()).all()
    
    # Enriquecer com informações de tipo e realização
    resultado = []
    for meta in metas:
        meta_dict = MetaOut.model_validate(meta).model_dump()
        
        # Adicionar nome do tipo
        if meta.tipo_lancamento_id:
            tipo = db.query(TipoLancamento).filter(TipoLancamento.id == meta.tipo_lancamento_id).first()
            if tipo:
                meta_dict["tipo_nome"] = tipo.nome
                meta_dict["tipo_natureza"] = tipo.natureza
        else:
            meta_dict["tipo_nome"] = "Geral"
            meta_dict["tipo_natureza"] = "ambos"
        
        # Calcular valor realizado
        valor_realizado = calcular_valor_realizado(
            db, meta.ano, meta.mes, meta.tipo_lancamento_id
        )
        meta_dict["valor_realizado"] = valor_realizado
        
        # Calcular percentual e status
        if float(meta.valor_planejado) > 0:
            percentual = (valor_realizado / float(meta.valor_planejado)) * 100
            meta_dict["percentual_realizado"] = round(percentual, 2)
            
            if percentual <= 90:
                meta_dict["status"] = "dentro"
            elif percentual <= 100:
                meta_dict["status"] = "atencao"
            else:
                meta_dict["status"] = "excedido"
        else:
            meta_dict["percentual_realizado"] = 0
            meta_dict["status"] = "dentro"
        
        resultado.append(meta_dict)
    
    return resultado

@app.get("/api/metas/{meta_id}")
async def obter_meta(meta_id: int, db: Session = Depends(get_db)):
    """Obtém uma meta específica com progresso"""
    meta = db.query(Meta).filter(Meta.id == meta_id).first()
    
    if not meta:
        raise HTTPException(status_code=404, detail="Meta não encontrada")
    
    meta_dict = MetaOut.model_validate(meta).model_dump()
    
    # Adicionar informações do tipo
    if meta.tipo_lancamento_id:
        tipo = db.query(TipoLancamento).filter(TipoLancamento.id == meta.tipo_lancamento_id).first()
        if tipo:
            meta_dict["tipo_nome"] = tipo.nome
            meta_dict["tipo_natureza"] = tipo.natureza
    
    # Calcular realização
    valor_realizado = calcular_valor_realizado(
        db, meta.ano, meta.mes, meta.tipo_lancamento_id
    )
    meta_dict["valor_realizado"] = valor_realizado
    
    if float(meta.valor_planejado) > 0:
        percentual = (valor_realizado / float(meta.valor_planejado)) * 100
        meta_dict["percentual_realizado"] = round(percentual, 2)
        
        if percentual <= 90:
            meta_dict["status"] = "dentro"
        elif percentual <= 100:
            meta_dict["status"] = "atencao"
        else:
            meta_dict["status"] = "excedido"
    
    return meta_dict

@app.put("/api/metas/{meta_id}")
async def atualizar_meta(meta_id: int, meta: MetaIn, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Atualiza uma meta existente"""
    from datetime import date as dt_date
    
    meta_db = db.query(Meta).filter(Meta.id == meta_id).first()
    
    if not meta_db:
        raise HTTPException(status_code=404, detail="Meta não encontrada")
    
    # Verificar duplicação (exceto a própria meta)
    meta_duplicada = db.query(Meta).filter(
        Meta.id != meta_id,
        Meta.ano == meta.ano,
        Meta.mes == meta.mes,
        Meta.tipo_lancamento_id == meta.tipo_lancamento_id
    ).first()
    
    if meta_duplicada:
        raise HTTPException(
            status_code=400,
            detail="Já existe outra meta para este período e tipo"
        )
    
    meta_db.ano = meta.ano
    meta_db.mes = meta.mes
    meta_db.tipo_lancamento_id = meta.tipo_lancamento_id
    meta_db.valor_planejado = meta.valor_planejado
    meta_db.descricao = meta.descricao
    meta_db.updated_at = dt_date.today()
    
    db.commit()
    db.refresh(meta_db)
    
    return MetaOut.model_validate(meta_db)

@app.delete("/api/metas/{meta_id}")
async def deletar_meta(meta_id: int, current_user: User = Depends(ensure_subscription), db: Session = Depends(get_db)):
    """Deleta uma meta"""
    meta = db.query(Meta).filter(Meta.id == meta_id).first()
    
    if not meta:
        raise HTTPException(status_code=404, detail="Meta não encontrada")
    
    db.delete(meta)
    db.commit()
    
    return {"success": True, "message": "Meta deletada com sucesso"}

@app.get("/api/metas/progresso/{ano}/{mes}")
async def obter_progresso_mes(ano: int, mes: int, db: Session = Depends(get_db)):
    """Obtém progresso geral do mês comparando metas vs realizado"""
    metas = db.query(Meta).filter(Meta.ano == ano, Meta.mes == mes).all()
    
    if not metas:
        return {
            "ano": ano,
            "mes": mes,
            "tem_metas": False,
            "total_planejado": 0,
            "total_realizado": 0,
            "percentual_geral": 0,
            "metas": []
        }
    
    total_planejado = 0
    total_realizado = 0
    metas_detalhes = []
    
    for meta in metas:
        valor_realizado = calcular_valor_realizado(
            db, meta.ano, meta.mes, meta.tipo_lancamento_id
        )
        
        planejado = float(meta.valor_planejado)
        total_planejado += planejado
        total_realizado += valor_realizado
        
        percentual = (valor_realizado / planejado * 100) if planejado > 0 else 0
        
        tipo_nome = "Geral"
        tipo_natureza = "ambos"
        if meta.tipo_lancamento_id:
            tipo = db.query(TipoLancamento).filter(TipoLancamento.id == meta.tipo_lancamento_id).first()
            if tipo:
                tipo_nome = tipo.nome
                tipo_natureza = tipo.natureza
        
        metas_detalhes.append({
            "id": meta.id,
            "tipo_nome": tipo_nome,
            "tipo_natureza": tipo_natureza,
            "valor_planejado": planejado,
            "valor_realizado": valor_realizado,
            "percentual": round(percentual, 2),
            "status": "dentro" if percentual <= 90 else "atencao" if percentual <= 100 else "excedido"
        })
    
    percentual_geral = (total_realizado / total_planejado * 100) if total_planejado > 0 else 0
    
    return {
        "ano": ano,
        "mes": mes,
        "tem_metas": True,
        "total_planejado": total_planejado,
        "total_realizado": total_realizado,
        "percentual_geral": round(percentual_geral, 2),
        "status_geral": "dentro" if percentual_geral <= 90 else "atencao" if percentual_geral <= 100 else "excedido",
        "metas": metas_detalhes
    }

def calcular_valor_realizado(db: Session, ano: int, mes: int, tipo_lancamento_id: Optional[int]) -> float:
    """Calcula valor realizado (pago) para um período e tipo"""
    from sqlalchemy import extract
    
    query = db.query(func.sum(Parcela.valor_pago)).join(
        Lancamento, Parcela.lancamento_id == Lancamento.id
    ).filter(
        Parcela.paga == 1,
        extract('year', Parcela.data_pagamento) == ano,
        extract('month', Parcela.data_pagamento) == mes
    )
    
    if tipo_lancamento_id:
        # Meta específica para um tipo
        query = query.filter(Lancamento.tipo_lancamento_id == tipo_lancamento_id)
        
        # Verificar se é despesa ou receita para considerar apenas um lado
        tipo = db.query(TipoLancamento).filter(TipoLancamento.id == tipo_lancamento_id).first()
        if tipo and tipo.natureza == "despesa":
            query = query.filter(Lancamento.tipo == "despesa")
        elif tipo and tipo.natureza == "receita":
            query = query.filter(Lancamento.tipo == "receita")
    
    total = query.scalar() or 0
    return float(total)

@app.get("/health")
def health():
    return {"status": "ok"}
